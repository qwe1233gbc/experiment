# -*- coding: utf-8 -*-
"""40题知识覆盖审计（81源版）生成脚本：只读引用Gold主表人工结论，不修改Gold/任何冻结文件。
输出：40题知识覆盖审计_81源版_20260820.xlsx + 40题知识覆盖审计总结_81源版_20260820.md
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REC = r"E:\实验文件整理_按论文逻辑\04_知识库构建\81源知识库入库执行记录_20260820"
XLSX = REC + r"\40题知识覆盖审计_81源版_20260820.xlsx"
MD = REC + r"\40题知识覆盖审计总结_81源版_20260820.md"

S = {
    "cons": "review-eia-construction-content-completeness",
    "ss": "calculate-eia-pollutant-source-strength",
    "coef": "review-eia-pollutant-coefficient-applicability",
    "air": "calculate-eia-exhaust-capture-airflow",
    "design": "review-eia-exhaust-design-airflow",
    "capeff": "review-eia-exhaust-capture-efficiency",
    "ac": "review-eia-activated-carbon-parameters",
    "hw": "review-eia-hazardous-waste-identification",
    "vt": "review-eia-vocs-total-control",
}
N_CN = {
    "cons": "建设内容完整性", "ss": "污染源强定量核算", "coef": "产污系数适用性",
    "air": "废气收集形式与理论排气量", "design": "废气设计风量", "capeff": "废气收集效率",
    "ac": "活性炭治理设施参数", "hw": "危险废物识别", "vt": "VOCs总量控制与一致性",
}

A = {
 "PL006_Construction_Q01": ("建设内容完整性", S["cons"], "工程分析内容齐全，无明显缺漏",
  "#5_建设内容编制完整性指南.md；HJ_884-2018.md", "B",
  "#31_环评报告表建设内容编制完整性清单（环办环评〔2020〕33号附件）",
  "工程组成五类框架可支撑方向判断；逐项清单仍未入库（建议补）"),
 "PL006_SourceStrength_Q01": ("污染源强定量核算", S["ss"], "源强核算存在不一致，需修正（燃烧废气速率与运行时间不复算）",
  "HJ_884-2018.md；#10_产污系数定量核算指南.md；#27_污染物源强产生收集处理排放闭合核算指南.md", "A",
  "燃气工业炉窑/天然气燃烧产污系数（建议#32，不影响本题不一致判定）",
  "#27新增产生→收集→治理→排放分配链条与kg/h×h/a一致性规则，90%收集×51%处理复算链可支撑"),
 "PL006_CaptureEfficiency_Q01": ("废气收集效率", S["capeff"], "90%收集效率合理，偏保守",
  "#19_废气收集效率判定指南.md", "A", "—", "#19取值表直接命中Gold判断"),
 "PL006_ActivatedCarbon_Q01": ("活性炭治理设施参数", S["ac"], "活性炭参数存在明显不合理和前后不一致，需修正",
  "HJ_2026-2013.md；#28_活性炭吸附治理参数核算审核指南.md", "A", "—",
  "#28新增更换周期口径一致性、碘值800mg/g、炭层300mm、15%/20%边界，Gold复算链条全部可支撑"),
 "PL007_Construction_Q01": ("建设内容完整性", S["cons"], "工程分析内容齐全",
  "#5_建设内容编制完整性指南.md；HJ_884-2018.md", "B",
  "#31_环评报告表建设内容编制完整性清单（环办环评〔2020〕33号附件）", "同PL006_Construction"),
 "PL007_SourceStrength_Q01": ("污染源强定量核算", S["ss"], "源强核算前后不一致，需修正",
  "HJ_884-2018.md；#10_产污系数定量核算指南.md；#25_涂料制造行业产污系数手册摘录_2641.md；#27_污染物源强产生收集处理排放闭合核算指南.md", "A", "—",
  "#25提供水性工业涂料2.00kg/t-产品系数，#27提供逐工序/三方一致性规则，正文vs源强表不一致可判定"),
 "PL007_DesignAirflow_Q01": ("废气设计风量", S["design"], "设计风量不合理，余量偏小（8.5%<10%下限）",
  "#18_废气收集风量与设计风量核算指南.md", "A", "—", "#18漏风10%-20%规则与Gold复算完全对应"),
 "PL007_CaptureEfficiency_Q01": ("废气收集效率", S["capeff"], "收集效率取值合理",
  "#19_废气收集效率判定指南.md", "A", "—", ""),
 "PL008_Construction_Q01": ("建设内容完整性", S["cons"], "工程分析内容齐全",
  "#5_建设内容编制完整性指南.md；HJ_884-2018.md", "B",
  "#31_环评报告表建设内容编制完整性清单（环办环评〔2020〕33号附件）", "同PL006_Construction"),
 "PL008_SourceStrength_Q01": ("污染源强定量核算", S["ss"], "源强核算存在问题，不能判定为正确（UV套用溶剂型系数）",
  "HJ_884-2018.md；#10_产污系数定量核算指南.md；#25_涂料制造行业产污系数手册摘录_2641.md；#27_污染物源强产生收集处理排放闭合核算指南.md", "A", "—",
  "#25提供UV/光固化无条目负证据与替代核算规则、2641水性系数，#27提供质量平衡闭合规则"),
 "PL008_DesignAirflow_Q01": ("废气设计风量", S["design"], "设计风量合理（16.6%位于10%-20%区间）",
  "#18_废气收集风量与设计风量核算指南.md", "A", "—", ""),
 "PL008_VOCSTotal_Q01": ("VOCs总量控制与一致性", S["vt"], "VOCs总量控制指标前后一致",
  "#29_VOCs排放核算与总量一致性审核指南.md；#27_污染物源强产生收集处理排放闭合核算指南.md；GB_37822-2019.md", "A", "—",
  "#29提供总排放=有组织+无组织明文与迁建新增量核算规则"),
 "PL009_Construction_Q01": ("建设内容完整性", S["cons"], "建设内容完整",
  "#5_建设内容编制完整性指南.md；HJ_884-2018.md", "B",
  "#31_环评报告表建设内容编制完整性清单（环办环评〔2020〕33号附件）", "同PL006_Construction"),
 "PL009_Coefficient_Q01": ("产污系数适用性", S["coef"], "产污系数部分不适用，需修正",
  "#9_塑料制品业产污系数_废气_官方手册.md；#10_产污系数定量核算指南.md；#26_废弃资源综合利用行业系数摘录_4220.md；#23_机械行业产污系数手册摘录_33-37_431-434.md", "A", "—",
  "#26提供4220干法破碎系数与‘自身边角料回用≠废塑料资源化’边界，#23提供下料5.30工艺边界（钻铣车削不适用）"),
 "PL009_DesignAirflow_Q01": ("废气设计风量", S["design"], "设计风量合理（21.4%属取整口径）",
  "#18_废气收集风量与设计风量核算指南.md", "A", "—", "裁量依据为#18规则本身"),
 "PL009_CaptureEfficiency_Q01": ("废气收集效率", S["capeff"], "收集效率总体合理，应统一表述",
  "#19_废气收集效率判定指南.md", "A", "—", ""),
 "PL010_SourceStrength_Q01": ("污染源强定量核算", S["ss"], "源强核算存在明显不一致，需修正",
  "HJ_884-2018.md；#10_产污系数定量核算指南.md；#23_机械行业产污系数手册摘录_33-37_431-434.md；#27_污染物源强产生收集处理排放闭合核算指南.md", "A", "—",
  "#23提供喷粉300kg/t、金属加工5.3kg/t口径，#27提供串联综合效率η总=1-(1-η1)(1-η2)与汇总一致性"),
 "PL010_Coefficient_Q01": ("产污系数适用性", S["coef"], "产污系数部分不适用，需修正",
  "#10_产污系数定量核算指南.md；#23_机械行业产污系数手册摘录_33-37_431-434.md", "A", "—",
  "#23提供喷粉300、固化VOCs 1.20、下料5.30条目及工艺边界（开料/冲压/机械加工不适用下料系数）"),
 "PL010_CaptureAirflow_Q01": ("废气收集形式与理论排气量", S["air"], "收集形式及理论排气量合理，计算可复核",
  "#17_废气收集形式及排气量计算指南.md；#18_废气收集风量与设计风量核算指南.md；#30_废气收集形式与排风量计算审核指南.md", "A", "—",
  "#30提供周长法Q=KPHV×3600（K=1.4工程参考）与换气次数法复算边界；60次/h按项目/工程依据核验，不作通用阈值"),
 "PL010_VOCSTotal_Q01": ("VOCs总量控制与一致性", S["vt"], "VOCs总量前后不一致，需修正",
  "#29_VOCs排放核算与总量一致性审核指南.md；GB_37822-2019.md", "A", "—",
  "#29提供总排放=有组织+无组织明文，0.194+0.047=0.241≠0.211判定有据"),
 "PL011_Construction_Q01": ("建设内容完整性", S["cons"], "建设内容完整",
  "#5_建设内容编制完整性指南.md；HJ_884-2018.md", "B",
  "#31_环评报告表建设内容编制完整性清单（环办环评〔2020〕33号附件）", "同PL006_Construction"),
 "PL011_CaptureAirflow_Q01": ("废气收集形式与理论排气量", S["air"], "收集形式及理论排气量合理",
  "#17_废气收集形式及排气量计算指南.md；#18_废气收集风量与设计风量核算指南.md；#30_废气收集形式与排风量计算审核指南.md", "A", "—",
  "#30提供周长法Q=1.4×P×H×V×3600工程参考公式，可复算4233.6m³/h"),
 "PL011_ActivatedCarbon_Q01": ("活性炭治理设施参数", S["ac"], "活性炭参数完整且核算合理",
  "HJ_2026-2013.md；#28_活性炭吸附治理参数核算审核指南.md", "A", "—",
  "#28提供15%（广东2023，版本路由）、理论需求=削减量/吸附比例、废活性炭=新鲜炭+吸附量、更换频次校核"),
 "PL011_HazardousWaste_Q01": ("危险废物识别", S["hw"], "危废识别完整，类别和代码合理",
  "EVID_HW_LIST_2025.md；EVID_HW_LIST_2021.md；EVID_GD_SOLID_WASTE_REG.md；GB_18597-2023.md", "A", "—", "两版名录均在库，代码核验充分"),
 "PL012_CaptureAirflow_Q01": ("废气收集形式与理论排气量", S["air"], "收集形式及理论排气量合理",
  "#17_废气收集形式及排气量计算指南.md；#18_废气收集风量与设计风量核算指南.md；#30_废气收集形式与排风量计算审核指南.md", "A", "—",
  "#30提供周长法K=1.4公式与工程参考边界，可复算9072m³/h与10.2%余量口径"),
 "PL012_ActivatedCarbon_Q01": ("活性炭治理设施参数", S["ac"], "活性炭参数完整且合理",
  "HJ_2026-2013.md；#28_活性炭吸附治理参数核算审核指南.md", "A", "—",
  "#28提供15%吸附比例（广东2023）与理论需求=削减量/吸附比例公式，0.01/15%=0.066t/a链条有据"),
 "PL012_HazardousWaste_Q01": ("危险废物识别", S["hw"], "危废识别完整，类别和代码合理",
  "EVID_HW_LIST_2025.md；EVID_HW_LIST_2021.md；EVID_GD_SOLID_WASTE_REG.md", "A", "—", ""),
 "PL012_VOCSTotal_Q01": ("VOCs总量控制与一致性", S["vt"], "VOCs总量控制指标前后一致",
  "#29_VOCs排放核算与总量一致性审核指南.md；GB_37822-2019.md", "A", "—",
  "#29提供总排放=有组织+无组织明文，0.010+0.011=0.021判定有据"),
 "PL013_SourceStrength_Q01": ("污染源强定量核算", S["ss"], "源强核算存在前后不一致，需修正",
  "HJ_884-2018.md；#10_产污系数定量核算指南.md；#27_污染物源强产生收集处理排放闭合核算指南.md", "A", "—",
  "#27提供物料衡算E投用=Σ(Wi×WFi)（MSDS挥发比例核验依据）与三方汇总一致性规则，0.013/0.007两值裁定有据"),
 "PL013_Coefficient_Q01": ("产污系数适用性", S["coef"], "产污系数/核算依据选取合理",
  "#10_产污系数定量核算指南.md；HJ_884-2018.md；#24_电子电气行业产污系数手册摘录_38-40.md", "A", "—",
  "#24提供无铅锡条波峰焊4.134e-1、无铅锡膏回流焊3.638e-1 g/kg-焊料条目及含铅/无铅/焊料形态/工艺边界"),
 "PL013_ActivatedCarbon_Q01": ("活性炭治理设施参数", S["ac"], "活性炭参数总体合理，用量满足吸附需求",
  "HJ_2026-2013.md；#28_活性炭吸附治理参数核算审核指南.md", "A", "—",
  "#28确认20%不得作通用值，本题按报告自身20%口径核验（边界内处理）；理论需求与废活性炭闭合公式可复算"),
 "PL013_HazardousWaste_Q01": ("危险废物识别", S["hw"], "危废识别完整，类别和代码与产生环节基本匹配",
  "EVID_HW_LIST_2025.md；EVID_HW_LIST_2021.md", "A", "—", "HW08/HW49代码条目均在库"),
 "PL014_Coefficient_Q01": ("产污系数适用性", S["coef"], "产污系数部分适用，UV涂料核算依据不足",
  "#10_产污系数定量核算指南.md；#25_涂料制造行业产污系数手册摘录_2641.md", "A", "—",
  "#25提供2641水性工业涂料2.00kg/t-产品系数、UV/光固化无条目负证据与替代核算规则（0.5%地方系数需单独核验）"),
 "PL014_CaptureAirflow_Q01": ("废气收集形式与理论排气量", S["air"], "核算方法不合理，需重新计算",
  "#17_废气收集形式及排气量计算指南.md；#18_废气收集风量与设计风量核算指南.md；#30_废气收集形式与排风量计算审核指南.md", "A", "—",
  "#30明文：收集效率用于污染物质量分配，不得用收集效率折减理论换气量/设计风量；12次/h换气法按工程参考复核"),
 "PL014_DesignAirflow_Q01": ("废气设计风量", S["design"], "设计风量不合理，24200m³/h偏小",
  "#18_废气收集风量与设计风量核算指南.md", "A", "—", "复算基于报告自身换气量数据，#18规则充分"),
 "PL014_VOCSTotal_Q01": ("VOCs总量控制与一致性", S["vt"], "总量加总正确但前后不一致（80%/90%口径），需修正",
  "#29_VOCs排放核算与总量一致性审核指南.md；#27_污染物源强产生收集处理排放闭合核算指南.md；GB_37822-2019.md；三线一单_顺德管控单元准入清单.md", "A",
  "行政总量政策（区域替代/以新带老）具体文件值仍按项目时点另行路由",
  "#29提供总排放/迁建/新增/削减规则与收集效率联动一致性（80%/90%口径典型错误）；内部核算≠行政政策合规边界已明确"),
 "PL015_Coefficient_Q01": ("产污系数适用性", S["coef"], "产污系数部分适用，喷粉核算基准需修正或说明",
  "#10_产污系数定量核算指南.md；#23_机械行业产污系数手册摘录_33-37_431-434.md；#27_污染物源强产生收集处理排放闭合核算指南.md", "A",
  "燃气工业炉窑/天然气燃烧产污系数（建议#32，不影响本题基准判定）",
  "#23提供固化VOCs 1.20kg/t-原料条目与核算基准=原料用量，#27提供基准/口径一致性规则，213 vs 191.46基准裁定有据"),
 "PL015_CaptureAirflow_Q01": ("废气收集形式与理论排气量", S["air"], "收集形式及理论排气量合理",
  "#17_废气收集形式及排气量计算指南.md；#18_废气收集风量与设计风量核算指南.md", "A", "—", "13.4%余量位于10%-20%区间，#17/#18规则充分"),
 "PL015_CaptureEfficiency_Q01": ("废气收集效率", S["capeff"], "50%收集效率合理且保守",
  "#19_废气收集效率判定指南.md", "A", "—", ""),
 "PL015_HazardousWaste_Q01": ("危险废物识别", S["hw"], "危废识别完整，类别和代码合理",
  "EVID_HW_LIST_2025.md；EVID_HW_LIST_2021.md；GB_18597-2023.md", "A", "—",
  "HW17(336-064-17)、HW49(900-039-49)可查；GB18597-2023在库可核对旧版标准引用"),
}

# 读Gold主表提取Gold核心结论（只读）
GOLD_CORE = {}
try:
    from openpyxl import load_workbook
    wb = load_workbook(r"E:\实验文件整理_按论文逻辑\05_QA测试集与样本\61题人工审阅记录表_剩余40题人工Gold完善版_20260819.xlsx", read_only=True, data_only=True)
    ws = wb["61题人工审阅记录表"]
    for row in ws.iter_rows(values_only=True):
        qid = str(row[1]) if row[1] else ""
        if qid in A:
            gold = (row[8] or "").split("\n")[0] if row[8] else ""
            GOLD_CORE[qid] = gold
except Exception as e:
    print("warn: Gold读取失败，使用内置结论", e)

ROWS = []
for qid, (atype, skill, gold_inline, avail, grade, miss, note) in A.items():
    pid = qid.split("_")[0]
    gold = GOLD_CORE.get(qid, gold_inline)
    impact = "否"
    ROWS.append([qid, pid, atype, skill, gold, avail, grade, miss, impact, note])

order = [S["cons"], S["ss"], S["coef"], S["air"], S["design"], S["capeff"], S["ac"], S["hw"], S["vt"]]
order = [sk for sk in order if any(r[3] == sk for r in ROWS)]

# ---------------- XLSX ----------------
wb = Workbook()
ws = wb.active
ws.title = "40题知识覆盖审计"
HDR = ["question_id", "project_id", "audit_type", "skill_id", "Gold核心结论",
       "当前可用正式知识", "知识充分性A/B/C", "仍缺知识", "是否影响正式实验", "备注"]
ws.append(HDR)
thin = Side(style="thin", color="AAB4C5")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="D9E1F2")
hdr_font = Font(name="微软雅黑", size=9, bold=True, color="1F3864")
cell_font = Font(name="微软雅黑", size=9)
zebra = PatternFill("solid", fgColor="F2F5FB")
fill_a = PatternFill("solid", fgColor="C6EFCE")
fill_b = PatternFill("solid", fgColor="FFEB9C")
fill_c = PatternFill("solid", fgColor="FFC7CE")
for c in range(1, len(HDR) + 1):
    cell = ws.cell(1, c)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, row in enumerate(ROWS, start=2):
    ws.append(row)
    gf = {"A": fill_a, "B": fill_b, "C": fill_c}[row[6]]
    for c in range(1, len(HDR) + 1):
        cell = ws.cell(i, c)
        cell.font = Font(name="微软雅黑", size=9, bold=(c == 7))
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c == 7:
            cell.fill = gf
        elif i % 2 == 0:
            cell.fill = zebra
widths = [24, 9, 15, 42, 40, 46, 9, 40, 11, 46]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{len(ROWS)+1}"

ws2 = wb.create_sheet("汇总统计")
ws2["A1"] = "40题知识覆盖审计总统计（81源版·20260820）"
ws2["A1"].font = Font(name="微软雅黑", bold=True, size=12)
ws2["A3"] = "级别"; ws2["B3"] = "题数"
for c in ("A3", "B3"):
    ws2[c].fill = hdr_fill; ws2[c].font = hdr_font; ws2[c].border = border
nA = sum(1 for r in ROWS if r[6] == "A"); nB = sum(1 for r in ROWS if r[6] == "B"); nC = sum(1 for r in ROWS if r[6] == "C")
for i, (k, v) in enumerate([("A级（充分）", nA), ("B级（部分充分）", nB), ("C级（不足）", nC), ("合计", len(ROWS))], start=4):
    ws2.cell(i, 1, k); ws2.cell(i, 2, v)
    for c in (1, 2):
        ws2.cell(i, c).font = cell_font; ws2.cell(i, c).border = border
ws2["A9"] = "Skill"; ws2["B9"] = "题数"; ws2["C9"] = "A"; ws2["D9"] = "B"; ws2["E9"] = "C"
for c in ("A9", "B9", "C9", "D9", "E9"):
    ws2[c].fill = hdr_fill; ws2[c].font = hdr_font; ws2[c].border = border
for i, sk in enumerate(order, start=10):
    sub = [r for r in ROWS if r[3] == sk]
    ws2.cell(i, 1, N_CN[[k for k, v in S.items() if v == sk][0]] + "（" + sk + "）")
    ws2.cell(i, 2, len(sub))
    ws2.cell(i, 3, sum(1 for r in sub if r[6] == "A"))
    ws2.cell(i, 4, sum(1 for r in sub if r[6] == "B"))
    ws2.cell(i, 5, sum(1 for r in sub if r[6] == "C"))
    for c in range(1, 6):
        ws2.cell(i, c).font = cell_font; ws2.cell(i, c).border = border
base = 10 + len(order)
ws2.cell(base, 1, "合计"); ws2.cell(base, 2, len(ROWS)); ws2.cell(base, 3, nA); ws2.cell(base, 4, nB); ws2.cell(base, 5, nC)
for c in range(1, 6):
    ws2.cell(base, c).font = Font(name="微软雅黑", size=9, bold=True); ws2.cell(base, c).border = border
ws2.cell(base + 2, 1, "未达A级题目")
for i, r in enumerate([r for r in ROWS if r[6] != "A"], start=base + 3):
    ws2.cell(i, 1, r[0]); ws2.cell(i, 2, r[6]); ws2.cell(i, 3, r[7]); ws2.cell(i, 4, r[1])
    for c in range(1, 5):
        ws2.cell(i, c).font = cell_font; ws2.cell(i, c).border = border
ws2.column_dimensions["A"].width = 62
for col in "BCDE":
    ws2.column_dimensions[col].width = 12
wb.save(XLSX)
print("saved:", XLSX)
print("A/B/C =", nA, nB, nC, "total =", len(ROWS))

# ---------------- MD ----------------
def dist(skill):
    sub = [r for r in ROWS if r[3] == skill]
    return (len(sub), sum(1 for r in sub if r[6] == "A"), sum(1 for r in sub if r[6] == "B"), sum(1 for r in sub if r[6] == "C"))
lines = []
lines.append("# 40题知识覆盖审计总结（81源版·20260820）\n")
lines.append("## 审计口径")
lines.append("- 知识库版本：`正式实验RAG知识库_81源扩展版_20260820`（73源→81源；401→409父块；3438→3459子块）")
lines.append("- 对照Gold：`61题人工审阅记录表_剩余40题人工Gold完善版_20260819.xlsx`（PL006–PL015，共40题，只读）")
lines.append("- 分级标准：A＝正式知识足以支撑Gold判断；B＝部分充分；C＝不足。**未为提升A级而降低标准**。\n")
lines.append("## 总览")
lines.append(f"- A级（充分）：**{nA}题**")
lines.append(f"- B级（部分充分）：**{nB}题**")
lines.append(f"- C级（不足）：**{nC}题**")
lines.append(f"- 合计：{len(ROWS)}题\n")
lines.append("## 各Skill A/B/C分布\n")
lines.append("| Skill | 题数 | A | B | C |")
lines.append("|---|---:|---:|---:|---:|")
for sk in order:
    key = [k for k, v in S.items() if v == sk][0]
    t, a, b, c = dist(sk)
    lines.append(f"| {N_CN[key]}（{sk}） | {t} | {a} | {b} | {c} |")
lines.append(f"| 合计 | {len(ROWS)} | {nA} | {nB} | {nC} |\n")
notA = [r for r in ROWS if r[6] != "A"]
lines.append("## 仍未达到A级的题目\n")
if notA:
    lines.append("| question_id | project_id | 级别 | 仍缺知识 | 备注 |")
    lines.append("|---|---|---|---|---|")
    for r in notA:
        lines.append(f"| {r[0]} | {r[1]} | {r[6]} | {r[7]} | {r[9]} |")
    lines.append("")
    lines.append("### 未达到A级的真实原因")
    lines.append("- 全部为「建设内容完整性」类题目（5题）。Gold核对项比`#5_建设内容编制完整性指南.md`的五类工程组成框架更细（产品产能/原辅材料/设备/工艺流程/水平衡/工作制度逐项要求），需要《建设项目环境影响报告表（污染影响类）编制技术指南》（环办环评〔2020〕33号）附件的编制内容完整性清单（建议文件`#31`）。")
    lines.append("- `#31`不在本次#23–#30正式入库包范围内，属**建议补充项（非必须项）**，不阻塞正式实验；如后续补充，可将5题升为A级。\n")
else:
    lines.append("无。\n")
lines.append("## 与20260819旧版（73源）对比")
lines.append("- 旧版：A=13、B=25、C=2。")
lines.append("- 新版（81源）：A=35、B=5、C=0。")
lines.append("- 升级来源：#23–#30正式入库后，原先因「系数/公式/边界明文缺失」判为B/C的题目获得正式依据（详见xlsx逐题备注）；唯一保留的B级均为建设内容完整性（#31未入库）。\n")
lines.append("## 结论")
lines.append("- 40题中35题已由正式知识充分支撑（A）；5题建设内容完整性为B级，原因统一为`#31`编制内容完整性清单未入库（建议补）。")
lines.append("- 未出现需要降低标准才能升A的情况；行政总量政策（区域替代/以新带老等）与广东VOCs 2021/2023版本差异已通过`版本路由_广东VOCs核算方法2021与2023.md`（Skill路由层manual_policy_files）与#29「内部核算≠行政政策合规」分层明确。")
open(MD, "w", encoding="utf-8").write("\n".join(lines))
print("saved:", MD)
