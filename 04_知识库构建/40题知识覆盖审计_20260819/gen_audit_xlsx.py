# -*- coding: utf-8 -*-
"""生成40题知识覆盖审计_20260819.xlsx（只读引用Gold主表，不修改任何原文件）"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"E:\实验文件整理_按论文逻辑\04_知识库构建\40题知识覆盖审计_20260819\40题知识覆盖审计_20260819.xlsx"

S = {
    "cons": "review-eia-construction-content-completeness",
    "ss": "calculate-eia-pollutant-source-strength",
    "coef": "review-eia-pollutant-coefficient-applicability",
    "air": "calculate-eia-exhaust-capture-airflow",
    "design": "review-eia-exhaust-design-airflow",
    "capeff": "review-eia-exhaust-capture-efficiency",
    "ac": "review-eia-activated-carbon-parameters",
    "hw": "review-eia-hazardous-waste-identification",
    "vt": "review-eia-vocs-total-control",
}

F = {
    "5": "#5_建设内容编制完整性指南.md",
    "9": "#9_塑料制品业产污系数_废气_官方手册.md",
    "10": "#10_产污系数定量核算指南.md",
    "17": "#17_废气收集形式及排气量计算指南.md",
    "18": "#18_废气收集风量与设计风量核算指南.md",
    "19": "#19_废气收集效率判定指南.md",
    "hj884": "HJ_884-2018.md",
    "hj2026": "HJ_2026-2013.md",
    "hw25": "EVID_HW_LIST_2025.md",
    "hw21": "EVID_HW_LIST_2021.md",
    "gbsolid": "EVID_GD_SOLID_WASTE_REG.md",
    "swlaw": "EVID_SOLID_WASTE_LAW_2020.md",
    "gb18597": "GB_18597-2023.md",
    "gb37822": "GB_37822-2019.md",
    "hj298": "EVID_HJ298_2019.md",
}

W = {
    "27": "#27_污染物源强产生收集处理排放闭合核算指南.md",
    "23": "#23_机械行业产污系数手册摘录_33-37_431-434.md",
    "24": "#24_电子电气行业产污系数手册摘录_38-40.md",
    "25": "#25_涂料制造行业产污系数手册摘录_2641.md",
    "26": "#26_废弃资源综合利用行业系数摘录_4220.md",
    "28": "#28_活性炭吸附治理参数核算审核指南.md",
    "29": "#29_VOCs总量控制核算与一致性审核指南.md",
    "30": "#30_废气收集排风量计算方法指南.md",
    "31": "#31_环评报告表建设内容编制完整性清单.md",
    "32": "#32_天然气燃烧及燃气工业炉窑产污系数摘录.md",
}

SRC_HB = "《排放源统计调查产排污核算方法和系数手册》（生态环境部公告2021年第24号）"
SRC_GT = "GB/T 16758-2009《排风罩的分类及技术条件》；AQ/T 4274-2016《局部排风设施控制风速检测与评估技术规范》"
SRC_VOC = "《广东省工业源VOCs减排量核算方法（2023修订版）》及省VOCs总量管理相关文件（#10/#18/#19已引用但原文未入库）"
SRC_AC = "HJ 2026-2013条文说明/省厅活性炭吸附治理设施规范化管理文件（15%/20%动态吸附容量取值需人工确认正式出处）"
SRC_ZB = "环办环评〔2020〕33号《建设项目环境影响报告表（污染影响类）编制技术指南》附件"

# (question_id, audit_type, skill, gold核心结论, RAG可用文件, 能支撑的判断, 级别, 缺失类型, 建议文件, 官方来源, 必须, 备注)
ROWS = [
 ("PL006_Construction_Q01", "建设内容完整性", S["cons"], "建设内容完整，工程分析内容齐全",
  F["5"] + "；" + F["hj884"],
  "按#5工程组成五类框架（主体/辅助/公用/环保/依托）核对工程组成齐全性",
  "B", "编制内容完整性清单（产品产能/原辅材料/设备/工艺流程/水平衡/工作制度逐项要求）",
  W["31"], SRC_ZB, "否", "Gold核对项比#5框架更细，方向判断可支撑"),
 ("PL006_SourceStrength_Q01", "污染源强定量核算", S["ss"], "源强核算存在不一致，需修正",
  F["hj884"] + "；" + F["10"] + "；" + F["9"],
  "物料衡算法/产污系数法选择依据（HJ884）；VOCs含量×用量的衡算思路",
  "B", "核算公式链（产生→收集→处理→有组织/无组织/总排放分配）；天然气燃烧SO₂/NOx/颗粒物系数；kg/h与年运行小时一致性校核规则",
  W["27"] + "；" + W["32"], "HJ 884-2018细化编制；" + SRC_HB + "燃气锅炉/工业炉窑分册", "是", "Gold复算90%收集×51%处理分配链条无库内公式依据"),
 ("PL006_CaptureEfficiency_Q01", "废气收集效率", S["capeff"], "90%收集效率合理，偏保守",
  F["19"],
  "单层密闭负压参考值95%+取值不得高于上限规则，90%<95%判断",
  "A", "无", "—", "—", "否", "#19取值表直接命中Gold判断"),
 ("PL006_ActivatedCarbon_Q01", "活性炭治理设施参数", S["ac"], "活性炭参数存在明显不合理和前后不一致，需修正",
  F["hj2026"],
  "更换触发原则（6.3.3.5）、用量按动态吸附量确定（6.3.3.2）、蜂窝炭流速≤1.2m/s、BET≥750m²/g",
  "B", "更换周期定量计算式（10000/18000m³/h两口径181天/101天复算）；碘值800mg/g指标（库内无碘值要求）；停留时间要求值；炭层布置规则",
  W["28"], SRC_AC, "是", "Gold全部定量结论依赖更换周期公式与碘值指标，HJ2026原文不含"),
 ("PL007_Construction_Q01", "建设内容完整性", S["cons"], "工程分析内容齐全",
  F["5"] + "；" + F["hj884"],
  "工程组成五类框架；物料平衡/水平衡要求方向",
  "B", "编制内容完整性清单（同PL006）",
  W["31"], SRC_ZB, "否", ""),
 ("PL007_SourceStrength_Q01", "污染源强定量核算", S["ss"], "源强核算前后不一致，需修正",
  F["hj884"] + "；" + F["10"],
  "核算方法依据；系数匹配原则",
  "B", "2641水性工业涂料2.00kg/t-产品系数；逐工序汇总一致性校核规则（正文vs源强表vs附表）",
  W["25"] + "；" + W["27"], SRC_HB + "涂料制造分册；HJ 884-2018细化编制", "是", "2.00kg/t系数库内无出处，无法验证报告取值"),
 ("PL007_DesignAirflow_Q01", "废气设计风量", S["design"], "设计风量不合理，余量偏小",
  F["18"],
  "设计风量=支管和×(1+漏风系数)、漏风10-20%，8.5%<10%下限判断",
  "A", "无", "—", "—", "否", "#18规则与Gold复算完全对应"),
 ("PL007_CaptureEfficiency_Q01", "废气收集效率", S["capeff"], "收集效率取值合理",
  F["19"],
  "设备排气口直连95%上限、单层密闭负压95%，90%/95%均不超限",
  "A", "无", "—", "—", "否", ""),
 ("PL008_Construction_Q01", "建设内容完整性", S["cons"], "工程分析内容齐全",
  F["5"] + "；" + F["hj884"],
  "工程组成框架含改扩建依托工程要求（#5要点5）",
  "B", "编制内容完整性清单；迁扩建前后工程变化说明要求",
  W["31"], SRC_ZB, "否", ""),
 ("PL008_SourceStrength_Q01", "污染源强定量核算", S["ss"], "源强核算存在问题，不能判定为正确",
  F["hj884"] + "；" + F["10"],
  "方法依据；UV涂料不得套用溶剂型系数的方向性判断（匹配原则）",
  "B", "UV涂料核算依据（系数手册无UV条目的事实+替代核算规则）；质量平衡逐项闭合校核规则；2641水性系数",
  W["25"] + "；" + W["27"], SRC_HB + "涂料制造分册；HJ 884-2018细化编制", "是", "Gold要求删除/复核溶剂型10kg/t系数，需手册边界证据"),
 ("PL008_DesignAirflow_Q01", "废气设计风量", S["design"], "设计风量合理",
  F["18"],
  "16.6%余量位于10-20%区间判断",
  "A", "无", "—", "—", "否", ""),
 ("PL008_VOCSTotal_Q01", "VOCs总量控制与一致性", S["vt"], "VOCs总量控制指标前后一致",
  F["gb37822"] + "（背景）；" + F["18"] + "/" + F["19"] + "（间接）",
  "VOCs分类收集与无组织管控背景；报告内部数据算术核对",
  "B", "总排放量=有组织+无组织的明文核算规则；迁建项目新增量核算规则（迁建前/后对比）",
  W["29"], SRC_VOC, "是", "核心加总规则未在正式RAG中找到明文"),
 ("PL009_Construction_Q01", "建设内容完整性", S["cons"], "建设内容完整",
  F["5"] + "；" + F["hj884"],
  "工程组成框架（按楼层布置的工程划分核对）",
  "B", "编制内容完整性清单",
  W["31"], SRC_ZB, "否", ""),
 ("PL009_Coefficient_Q01", "产污系数适用性", S["coef"], "产污系数部分不适用，需修正",
  F["9"] + "；" + F["10"],
  "2929注塑2.70kg/t-产品系数正式在库（#9手册2929系数表），注塑项适用判断；系数匹配原则",
  "B", "4220废塑料破碎系数及适用边界（自身边角料回用≠废塑料资源化）；机械行业锯床/砂轮/切割机5.30kg/t系数的工艺边界",
  W["26"] + "；" + W["23"], SRC_HB + "42废弃资源综合利用分册；33-37,431-434机械行业分册", "是", "候选索引有4220（375/425/450g/t）与机械下料5.30条目，但均为待官方核验状态"),
 ("PL009_DesignAirflow_Q01", "废气设计风量", S["design"], "设计风量合理",
  F["18"],
  "21.4%略超20%上界的裁量判断（10-20%规则+取整口径）",
  "A", "无", "—", "—", "否", "裁量依据为#18规则本身"),
 ("PL009_CaptureEfficiency_Q01", "废气收集效率", S["capeff"], "收集效率总体合理，应统一表述",
  F["19"],
  "包围式(≥0.5m/s)上限80%，50%保守；密闭收集85%未超限",
  "A", "无", "—", "—", "否", ""),
 ("PL010_SourceStrength_Q01", "污染源强定量核算", S["ss"], "源强核算存在明显不一致，需修正",
  F["hj884"] + "；" + F["10"],
  "方法依据；产生量=系数×用量复算思路",
  "B", "喷粉300kg/t-原料系数（库与候选索引均无，需从机械行业手册原文定位）；串联除尘综合效率公式（旋风60%+滤筒80%→92%）；金属加工5.3kg/t复算口径",
  W["23"] + "；" + W["27"], SRC_HB + "机械行业分册；HJ 884-2018细化编制", "是", "85%收集+两级除尘92%综合效率的复算链无库内公式"),
 ("PL010_Coefficient_Q01", "产污系数适用性", S["coef"], "产污系数部分不适用，需修正",
  F["10"],
  "系数须与行业/工艺/核算基准匹配的原则",
  "B", "喷粉300kg/t、固化VOCs 1.2kg/t、下料5.3kg/t的官方手册条目及适用工艺边界",
  W["23"], SRC_HB + "机械行业分册（涂装/下料/机加工表）", "是", "候选索引有涂装表VOCs 1.2kg/t及下料5.30条目（待核验），喷粉300kg/t连候选索引亦无"),
 ("PL010_CaptureAirflow_Q01", "废气收集形式与理论排气量", S["air"], "收集形式及理论排气量合理，计算可复核",
  F["17"] + "；" + F["18"],
  "收集形式分类与控制风速要求；设计风量=支管和×(1+漏风)汇总闭合",
  "B", "周长法公式L=1.4×P×h×Vk×3600（库内仅有罩口面积法Q=3600×F×v）；换气次数法（喷粉房60次/h）取值规则；GB/T 16758原文（GB37822引用但未入库）",
  W["30"], SRC_GT, "是", "Gold复算全部使用周长法与换气次数法，两公式均不在正式RAG"),
 ("PL010_VOCSTotal_Q01", "VOCs总量控制与一致性", S["vt"], "VOCs总量前后不一致，需修正",
  F["gb37822"] + "（背景）",
  "报告内部分项数据核对",
  "B", "总排放量=有组织+无组织明文规则（0.194+0.047=0.241≠0.211判断依据）",
  W["29"], SRC_VOC, "是", "核心规则未在正式RAG中找到明文"),
 ("PL011_Construction_Q01", "建设内容完整性", S["cons"], "建设内容完整",
  F["5"] + "；" + F["hj884"],
  "产品—原辅材料—设备—工艺—产污环节闭合链条方向核对",
  "B", "编制内容完整性清单",
  W["31"], SRC_ZB, "否", ""),
 ("PL011_CaptureAirflow_Q01", "废气收集形式与理论排气量", S["air"], "收集形式及理论排气量合理",
  F["17"] + "；" + F["18"],
  "包围型集气罩形式匹配；20%设计余量符合10-20%规则",
  "B", "周长法Q=K×P×H×Vx×3600（K=1.4）公式及K取值依据",
  W["30"], SRC_GT, "是", "Gold复算4233.6m³/h直接使用K=1.4公式"),
 ("PL011_ActivatedCarbon_Q01", "活性炭治理设施参数", S["ac"], "活性炭参数完整且核算合理",
  F["hj2026"],
  "用量与动态吸附量挂钩的定性规则；流速限值",
  "B", "15%动态吸附容量参考取值；理论活性炭需求=削减量/动态吸附容量公式（2.4149t/a复算）；废活性炭=新鲜炭+吸附VOCs量规则（3.9623t/a闭合）；更换频次与用量校核方法",
  W["28"], SRC_AC, "是", "Gold定量链条（15%→2.4149→3.6→3.9623）无库内依据"),
 ("PL011_HazardousWaste_Q01", "危险废物识别", S["hw"], "危废识别完整，类别和代码合理",
  F["hw25"] + "；" + F["hw21"] + "；" + F["gbsolid"] + "；" + F["gb18597"],
  "HW49(900-041-49含油抹布)、HW49(900-039-49废活性炭)条目均在2025/2021名录中可查；代码与产生环节匹配核对",
  "A", "无", "—", "—", "否", "两版名录均在库，代码核验充分"),
 ("PL012_CaptureAirflow_Q01", "废气收集形式与理论排气量", S["air"], "收集形式及理论排气量合理",
  F["17"] + "；" + F["18"],
  "集气罩收集形式；最终设计风量10.2%余量达标判断",
  "B", "周长法K=1.4公式（9072m³/h复算）；管道漏风率与设计余量口径关系",
  W["30"], SRC_GT, "是", "Gold复算使用K=1.4公式且区分8%管道漏风与10%设计余量口径"),
 ("PL012_ActivatedCarbon_Q01", "活性炭治理设施参数", S["ac"], "活性炭参数完整且合理",
  F["hj2026"],
  "过滤风速/停留时间几何复算可校核（装置尺寸在报告内）；蜂窝炭流速限值",
  "B", "15%吸附比例取值；理论需求0.066t/a核算公式；年活性炭量与需求比较规则",
  W["28"], SRC_AC, "是", "0.01/15%=0.066t/a链条无库内依据"),
 ("PL012_HazardousWaste_Q01", "危险废物识别", S["hw"], "危废识别完整，类别和代码合理",
  F["hw25"] + "；" + F["hw21"] + "；" + F["gbsolid"],
  "HW49(900-041-49)、HW49(900-039-49)条目可查",
  "A", "无", "—", "—", "否", ""),
 ("PL012_VOCSTotal_Q01", "VOCs总量控制与一致性", S["vt"], "VOCs总量控制指标前后一致",
  F["gb37822"] + "（背景）",
  "报告内部0.010+0.011=0.021算术核对",
  "B", "总排放量=有组织+无组织明文规则；总量控制指标申报值核定规则",
  W["29"], SRC_VOC, "是", "核心规则未在正式RAG中找到明文"),
 ("PL013_SourceStrength_Q01", "污染源强定量核算", S["ss"], "源强核算存在前后不一致，需修正",
  F["hj884"] + "；" + F["10"],
  "MSDS挥发组分物料衡算法依据（HJ884物料衡算定义）；数值一致性核查方向",
  "B", "MSDS组分比例核验规则（锡膏11%等取值可追溯性）；回流焊0.013/0.007两值裁定所需的汇总一致性规则",
  W["27"], "HJ 884-2018细化编制（物料衡算数据可追溯性要求）", "是", "衡算方法有依据，但组分比例核验与汇总裁定规则缺"),
 ("PL013_Coefficient_Q01", "产污系数适用性", S["coef"], "产污系数/核算依据选取合理",
  F["10"] + "；" + F["hj884"],
  "系数与工艺对应原则；物料衡算属可追溯核算方法的定性依据",
  "B", "电子电气行业无铅焊料波峰焊4.134×10⁻¹g/kg-焊料、回流焊3.638×10⁻¹g/kg-焊料手册条目（库与候选索引均无，需从38-40分册定位）",
  W["24"], SRC_HB + "电子电气行业分册（38-40）", "是", "两焊接系数值无法在正式RAG中验证"),
 ("PL013_ActivatedCarbon_Q01", "活性炭治理设施参数", S["ac"], "活性炭参数总体合理",
  F["hj2026"],
  "二级蜂窝炭碘值方向性要求（库内为BET≥750m²/g，无碘值指标）；用量与削减量校核原则",
  "B", "20%吸附能力取值；理论需求=0.592/20%=2.96t/a公式；废活性炭5.152=4.56+0.592闭合规则；碘值≥800mg/g指标",
  W["28"], SRC_AC, "是", "碘值800mg/g与20%取值均不在HJ2026原文"),
 ("PL013_HazardousWaste_Q01", "危险废物识别", S["hw"], "危废识别完整，类别和代码匹配",
  F["hw25"] + "；" + F["hw21"],
  "HW08(900-249-08)、HW49(900-041-49)、HW49(900-039-49)条目均可查",
  "A", "无", "—", "—", "否", ""),
 ("PL014_Coefficient_Q01", "产污系数适用性", S["coef"], "产污系数部分适用，UV涂料核算依据不足",
  F["10"] + "（仅匹配原则）",
  "系数须与行业/产品/工艺/基准匹配的原则",
  "C", "2641涂料制造水性工业涂料2.00kg/t-产品系数（库与候选索引均无）；UV涂料系数手册无条目的事实依据；UV涂料替代核算方法规则",
  W["25"], SRC_HB + "涂料制造分册（2641）", "是", "本题两个核心判断（2641适用、UV无手册条目）均无正式知识支撑"),
 ("PL014_CaptureAirflow_Q01", "废气收集形式与理论排气量", S["air"], "核算方法不合理，需重新计算",
  F["17"] + "；" + F["18"],
  "密闭/外部收集形式分类；设计风量=支管和×(1+漏风)规则（间接支撑不得折减）",
  "B", "换气次数法（空间体积×12次/h）取值规则；收集效率不得用于折减设计风量的明文禁令",
  W["30"], SRC_GT, "是", "Gold判断核心为折减禁令，正式RAG无明文"),
 ("PL014_DesignAirflow_Q01", "废气设计风量", S["design"], "设计风量不合理，24200m³/h偏小",
  F["18"],
  "设计风量构成规则+10-20%余量；24200<26028（报告未折减理论需求）判断",
  "A", "无", "—", "—", "否", "复算基于报告自身换气量数据，#18规则充分"),
 ("PL014_VOCSTotal_Q01", "VOCs总量控制与一致性", S["vt"], "总量加总正确但前后不一致，需修正",
  F["gb37822"] + "（背景）；三线一单_顺德管控单元准入清单.md（VOCs综合管控方向）",
  "报告内三组算术关系核对",
  "C", "总排放量=有组织+无组织明文；迁建前/迁建后/新增量/削减量核算规则；以新带老与区域替代要求；收集效率与总量联动一致性规则（80%/90%口径）",
  W["29"], SRC_VOC + "；总量管理制度文件", "是", "本题所需总量管理政策框架在正式RAG中基本缺失"),
 ("PL015_Coefficient_Q01", "产污系数适用性", S["coef"], "产污系数部分适用，喷粉基准需修正",
  F["10"] + "；#9_塑料制品业产污系数_废气_速查.md（核算口径提示）",
  "系数与工艺对应原则；核算基准一致性提示",
  "B", "固化VOCs 1.20kg/t-原料条目（候选索引有1.2kg/t待核验）；燃气工业炉窑产污系数；喷粉核算基准（总用量vs扣除回收后用量）规则",
  W["23"] + "；" + W["32"], SRC_HB + "机械行业分册；燃气工业炉窑分册", "是", "213t/a与191.46t/a基准裁定需核算基准规则"),
 ("PL015_CaptureAirflow_Q01", "废气收集形式与理论排气量", S["air"], "收集形式及理论排气量合理",
  F["17"] + "；" + F["18"],
  "集气罩收集形式匹配；13.4%设计余量位于10-20%区间判断",
  "A", "无", "—", "—", "否", "风量数据取自报告，#17/#18规则充分"),
 ("PL015_CaptureEfficiency_Q01", "废气收集效率", S["capeff"], "50%收集效率合理且保守",
  F["19"],
  "包围式(≥0.5m/s)上限80%，50%明显低于上限",
  "A", "无", "—", "—", "否", ""),
 ("PL015_HazardousWaste_Q01", "危险废物识别", S["hw"], "危废识别完整，类别和代码合理",
  F["hw25"] + "；" + F["hw21"] + "；" + F["gb18597"],
  "HW17(336-064-17)、HW49(900-039-49)条目可查；GB18597-2023在库可核对旧版标准引用问题",
  "A", "无", "—", "—", "否", ""),
]

HDR = ["question_id", "project_id", "audit_type", "skill_id", "人工Gold核心结论", "当前正式RAG可用知识文件",
       "当前知识能支撑的判断", "知识充分性：A/B/C", "缺失知识类型", "建议补充的正式知识文件", "建议补充的官方来源",
       "是否必须补齐后再跑正式实验", "备注"]

wb = Workbook()
ws = wb.active
ws.title = "40题知识覆盖审计"

thin = Side(style="thin", color="D9DEE7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
cell_font = Font(name="微软雅黑", size=9)
fill_a = PatternFill("solid", fgColor="E2EFDA")
fill_b = PatternFill("solid", fgColor="FFF2CC")
fill_c = PatternFill("solid", fgColor="FCE4E4")
zebra = PatternFill("solid", fgColor="F7F9FC")

ws.append(HDR)
for c in range(1, len(HDR) + 1):
    cell = ws.cell(1, c)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for i, r in enumerate(ROWS, start=2):
    qid, atype, skill, gold, rag, support, grade, miss, sug, src, must, note = r
    pid = qid.split("_")[0]
    row = [qid, pid, atype, skill, gold, rag, support, grade, miss, sug, src, must, note]
    ws.append(row)
    gf = {"A": fill_a, "B": fill_b, "C": fill_c}[grade]
    for c in range(1, len(HDR) + 1):
        cell = ws.cell(i, c)
        cell.font = Font(name="微软雅黑", size=9, bold=(c == 8))
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c == 8:
            cell.fill = gf
        elif i % 2 == 0:
            cell.fill = zebra

widths = [26, 8, 15, 34, 22, 34, 34, 9, 36, 30, 34, 12, 26]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:M{len(ROWS)+1}"

ws2 = wb.create_sheet("汇总统计")
ws2["A1"] = "40题知识覆盖审计总统计（20260819）"
ws2["A1"].font = Font(name="微软雅黑", bold=True, size=12)
stats = [
    ("A级（充分）", sum(1 for r in ROWS if r[6] == "A")),
    ("B级（部分充分）", sum(1 for r in ROWS if r[6] == "B")),
    ("C级（不足）", sum(1 for r in ROWS if r[6] == "C")),
    ("合计", len(ROWS)),
]
ws2["A3"] = "级别"; ws2["B3"] = "题数"
for c in ("A3", "B3"):
    ws2[c].fill = hdr_fill; ws2[c].font = hdr_font; ws2[c].border = border
for i, (k, v) in enumerate(stats, start=4):
    ws2.cell(i, 1, k); ws2.cell(i, 2, v)
    for c in (1, 2):
        ws2.cell(i, c).font = cell_font; ws2.cell(i, c).border = border

order = [S["cons"], S["ss"], S["coef"], S["air"], S["design"], S["capeff"], S["ac"], S["hw"], S["vt"]]
names = {S["cons"]: "建设内容完整性", S["ss"]: "污染源强定量核算", S["coef"]: "产污系数适用性",
         S["air"]: "废气收集形式与理论排气量", S["design"]: "废气设计风量", S["capeff"]: "废气收集效率",
         S["ac"]: "活性炭治理设施参数", S["hw"]: "危险废物识别", S["vt"]: "VOCs总量控制与一致性"}
ws2["A9"] = "Skill"; ws2["B9"] = "题数"; ws2["C9"] = "A"; ws2["D9"] = "B"; ws2["E9"] = "C"
for c in ("A9", "B9", "C9", "D9", "E9"):
    ws2[c].fill = hdr_fill; ws2[c].font = hdr_font; ws2[c].border = border
for i, sk in enumerate(order, start=10):
    sub = [r for r in ROWS if r[2] == sk]
    ws2.cell(i, 1, names[sk] + "（" + sk + "）")
    ws2.cell(i, 2, len(sub))
    ws2.cell(i, 3, sum(1 for r in sub if r[6] == "A"))
    ws2.cell(i, 4, sum(1 for r in sub if r[6] == "B"))
    ws2.cell(i, 5, sum(1 for r in sub if r[6] == "C"))
    for c in range(1, 6):
        ws2.cell(i, c).font = cell_font; ws2.cell(i, c).border = border

base = 10 + len(order)
ws2.cell(base, 1, "合计"); ws2.cell(base, 2, len(ROWS))
ws2.cell(base, 3, sum(1 for r in ROWS if r[6] == "A"))
ws2.cell(base, 4, sum(1 for r in ROWS if r[6] == "B"))
ws2.cell(base, 5, sum(1 for r in ROWS if r[6] == "C"))
for c in range(1, 6):
    ws2.cell(base, c).font = Font(name="微软雅黑", size=9, bold=True)
    ws2.cell(base, c).border = Border(left=thin, right=thin, bottom=thin, top=Side(style="medium", color="AAB4C5"))

ws2.cell(base + 2, 1, "正式实验前必须补的知识文件数量")
ws2.cell(base + 2, 2, 8)
ws2.cell(base + 3, 1, "建议补充但不是必须的知识文件数量")
ws2.cell(base + 3, 2, 2)
for rr in (base + 2, base + 3):
    for c in (1, 2):
        ws2.cell(rr, c).font = Font(name="微软雅黑", size=9, bold=True)
        ws2.cell(rr, c).fill = PatternFill("solid", fgColor="EAF2FF")
must_files = [
    "#23 机械行业产污系数手册摘录（33-37,431-434）", "#24 电子电气行业产污系数手册摘录（38-40）",
    "#25 涂料制造行业产污系数手册摘录（2641）", "#26 废弃资源综合利用行业系数摘录（4220）",
    "#27 污染物源强产生收集处理排放闭合核算指南", "#28 活性炭吸附治理参数核算审核指南",
    "#29 VOCs总量控制核算与一致性审核指南", "#30 废气收集排风量计算方法指南",
]
sug_files = ["#31 环评报告表建设内容编制完整性清单", "#32 天然气燃烧及燃气工业炉窑产污系数摘录"]
ws2.cell(base + 5, 1, "必须补清单").font = Font(name="微软雅黑", bold=True, size=10)
for i, m in enumerate(must_files):
    ws2.cell(base + 6 + i, 1, m).font = cell_font
ws2.cell(base + 6 + len(must_files) + 1, 1, "建议补清单").font = Font(name="微软雅黑", bold=True, size=10)
for i, m in enumerate(sug_files):
    ws2.cell(base + 7 + len(must_files) + i, 1, m).font = cell_font
ws2.column_dimensions["A"].width = 62
for col in "BCDE":
    ws2.column_dimensions[col].width = 10

wb.save(OUT)
print("saved:", OUT)
print("A/B/C =", sum(1 for r in ROWS if r[6]=="A"), sum(1 for r in ROWS if r[6]=="B"), sum(1 for r in ROWS if r[6]=="C"), "total =", len(ROWS))
