#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
为 Pilot 16 题生成 RAG 检索快照
使用 81 源正式知识库 + 混合检索（Dense + BM25 + Rerank）
"""
import sys
import os
import json
import hashlib
import openpyxl
from pathlib import Path

# 把自建脚本目录加入路径，导入HybridRetriever
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = Path(r"E:\实验文件整理_按论文逻辑")
sys.path.insert(0, str(REPO_ROOT / "自建脚本_统一管理"))

from hybrid_retrieval import HybridRetriever

# ========== 配置 ==========
EXP_DIR = REPO_ROOT / "实验"
QFILE = EXP_DIR / "02_evaluation_set" / "pilot16_questions.xlsx"
INDEX_DIR = EXP_DIR / "03_knowledge_base" / "retrieval_index"
OUTPUT_FILE = EXP_DIR / "03_knowledge_base" / "pilot16_rag_snapshot.jsonl"
MANIFEST_FILE = EXP_DIR / "03_knowledge_base" / "pilot16_rag_manifest.json"

DASHSCOPE_KEY = os.environ.get("DASHSCOPE_API_KEY", "")
FINAL_TOP_K = 5
DENSE_TOP_K = 40
SPARSE_TOP_K = 100
CANDIDATE_TOP_K = 50

# ========== 读取题目 ==========
def load_questions():
    wb = openpyxl.load_workbook(QFILE, data_only=True)
    ws = wb["01_题目清单"]
    headers = [c.value for c in ws[1]]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        q = dict(zip(headers, row))
        questions.append({
            "question_id": q["question_id"],
            "ep_category": q["EP类别"],
            "task_type": q.get("task_type", ""),
            "question": q["题目（question）"],
            "project": q.get("项目", ""),
        })
    return questions

# ========== 构造检索query ==========
def build_retrieval_query(q):
    """根据题目构造检索query。E1类用原题，E0类提取核心概念。"""
    question = q["question"]
    ep = q["ep_category"]
    
    # 直接用原题作为检索query（题目本身已经是自然语言问题）
    # 如果题目太长，截断前500字
    query = question.strip()
    if len(query) > 500:
        query = query[:500]
    
    return query

# ========== 主函数 ==========
def main():
    print("=" * 60)
    print("Pilot 16 题 RAG 检索快照生成")
    print("=" * 60)
    
    if not DASHSCOPE_KEY:
        print("❌ 缺少 DASHSCOPE_API_KEY 环境变量")
        sys.exit(1)
    
    # 加载题目
    questions = load_questions()
    print(f"\n加载题目：{len(questions)} 题")
    
    # 初始化检索器
    print(f"加载检索索引：{INDEX_DIR}")
    retriever = HybridRetriever(INDEX_DIR, DASHSCOPE_KEY)
    info = retriever.validate()
    print(f"  来源: {info['sources']}, 父块: {info['parents']}, 子块: {info['children']}")
    print(f"  向量维度: {info['dimensions']}, 模型: {info['embedding_model']}")
    
    # 断点续跑
    done_qids = set()
    results = []
    if OUTPUT_FILE.exists():
        with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    rec = json.loads(line)
                    results.append(rec)
                    done_qids.add(rec["question_id"])
        print(f"\n断点续跑：已完成 {len(done_qids)} 题")
    
    # 逐题检索
    for i, q in enumerate(questions):
        qid = q["question_id"]
        if qid in done_qids:
            print(f"[{i+1}/{len(questions)}] ⏭️  {qid} (已完成)")
            continue
        
        query = build_retrieval_query(q)
        print(f"[{i+1}/{len(questions)}] 🔍 {qid} (query长度: {len(query)})")
        
        try:
            result = retriever.retrieve(
                query=query,
                dense_top_k=DENSE_TOP_K,
                sparse_top_k=SPARSE_TOP_K,
                candidate_top_k=CANDIDATE_TOP_K,
                final_top_k=FINAL_TOP_K,
                use_rerank=True,
            )
            
            # 构造快照记录（每题top_k条）
            for j, hit in enumerate(result["hits"]):
                snapshot_record = {
                    "question_id": qid,
                    "ep_category": q["ep_category"],
                    "task_type": q["task_type"],
                    "rank": j + 1,
                    "child_id": hit["child_id"],
                    "parent_id": hit["parent_id"],
                    "source_id": hit["source_id"],
                    "source_file": hit["source_file"],
                    "retrieval_score": hit["score"],
                    "rerank_score": hit["score"],  # 因为用了rerank，score就是rerank分
                    "title": hit.get("source_file", ""),
                    "text": hit["parent_content"],  # 用父块全文作为证据
                    "child_text": hit["child_content"],  # 子块也保留
                    "retrieval_method": result["method"],
                    "retrieval_query_sha256": hashlib.sha256(query.encode("utf-8")).hexdigest(),
                    "snapshot_hash": hashlib.sha256(
                        f"{qid}|{hit['parent_id']}|{hit['score']}".encode("utf-8")
                    ).hexdigest(),
                }
                results.append(snapshot_record)
                
                # 立即写入
                with open(OUTPUT_FILE, "a", encoding="utf-8") as f:
                    f.write(json.dumps(snapshot_record, ensure_ascii=False) + "\n")
            
            print(f"       ✅ 找到 {len(result['hits'])} 条结果，{result['unique_parent_count']} 个不重复来源")
            
        except Exception as e:
            print(f"       ❌ 失败: {type(e).__name__}: {e}")
            # 失败也记录一条占位
            fail_record = {
                "question_id": qid,
                "ep_category": q["ep_category"],
                "task_type": q["task_type"],
                "rank": 0,
                "error": str(e),
                "status": "FAILED",
            }
            results.append(fail_record)
    
    # 统计
    success_qids = set()
    for r in results:
        if r.get("status") != "FAILED" and r.get("rank", 0) > 0:
            success_qids.add(r["question_id"])
    
    print(f"\n{'=' * 60}")
    print(f"完成！成功题目：{len(success_qids)} / {len(questions)}")
    print(f"总记录数：{len([r for r in results if r.get('rank', 0) > 0])} 条")
    print(f"输出文件：{OUTPUT_FILE}")
    
    # 生成manifest
    manifest = {
        "version": "pilot16_v1",
        "generated_at": __import__("datetime").datetime.now().isoformat(),
        "total_questions": len(questions),
        "successful_questions": len(success_qids),
        "total_records": len([r for r in results if r.get("rank", 0) > 0]),
        "final_top_k": FINAL_TOP_K,
        "index_source": "81源正式库_混合检索",
        "embedding_model": info["embedding_model"],
        "rerank_model": "gte-rerank-v2",
        "retrieval_method": "Dense+BM25 RRF -> gte-rerank-v2 -> source-dedup",
        "question_ids": sorted(success_qids),
        "output_file": str(OUTPUT_FILE.name),
    }
    with open(MANIFEST_FILE, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    print(f"Manifest：{MANIFEST_FILE}")
    
    # 按题目统计
    print(f"\n按EP类别统计：")
    ep_counts = {}
    for q in questions:
        ep = q["ep_category"]
        ep_counts.setdefault(ep, {"total": 0, "success": 0})
        ep_counts[ep]["total"] += 1
        if q["question_id"] in success_qids:
            ep_counts[ep]["success"] += 1
    
    for ep in sorted(ep_counts.keys()):
        c = ep_counts[ep]
        print(f"  {ep}: {c['success']}/{c['total']}")

if __name__ == "__main__":
    main()
