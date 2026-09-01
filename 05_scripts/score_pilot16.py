#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pilot 16 题自动评分脚本
支持多种题型：判断题、选择题、计算题、综合分析题
输出评分表和统计汇总
"""
import os
import json
import re
import openpyxl
from pathlib import Path

EXP_DIR = Path(r"E:\实验文件整理_按论文逻辑\实验")
RESULTS_FILE = EXP_DIR / "07_results" / "pilot16_raw_results.jsonl"
QFILE = EXP_DIR / "02_evaluation_set" / "pilot16_questions.xlsx"
OUTPUT_XLSX = EXP_DIR / "07_results" / "pilot16_scored.xlsx"
SUMMARY_MD = EXP_DIR / "07_results" / "pilot16_score_summary.md"

os.makedirs(EXP_DIR / "07_results", exist_ok=True)

def load_questions():
    wb = openpyxl.load_workbook(QFILE, data_only=True)
    ws = wb["01_题目清单"]
    headers = [c.value for c in ws[1]]
    questions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        q = dict(zip(headers, row))
        questions[q["question_id"]] = {
            "question_id": q["question_id"],
            "ep_category": q["EP类别"],
            "task_type": q.get("task_type", ""),
            "question": q["题目（question）"],
            "gold_answer": q.get("金标答案（gold）", ""),
        }
    return questions

def load_results():
    results = []
    if not RESULTS_FILE.exists():
        return results
    with open(RESULTS_FILE, "r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                results.append(json.loads(line))
    return results

def extract_conclusion(answer):
    """从模型回答中提取【结论】部分。"""
    if not answer:
        return ""
    
    # 尝试找【结论】标记
    patterns = [
        r"【结论】\s*\n?(.*?)(?=\n【|$)",
        r"结论[：:]\s*\n?(.*?)(?=\n\S+[：:]|\n[A-Z]|$)",
        r"^(.{0,100}?)(?=\n|$)",  # 取第一行
    ]
    
    for pattern in patterns:
        m = re.search(pattern, answer, re.DOTALL)
        if m:
            text = m.group(1).strip()
            if text and len(text) < 500:
                return text
    
    return answer[:200]

def score_answer(question, gold, answer):
    """
    自动评分：返回 (correctness, score, reason)
    correctness: CORRECT / INCORRECT / PARTIAL / UNKNOWN
    score: 0-1 分
    """
    if not answer or answer.strip() == "":
        return "UNKNOWN", 0, "空回答"
    
    conclusion = extract_conclusion(answer)
    gold_str = str(gold).strip()
    ans_str = str(answer).strip()
    concl_str = conclusion.strip()
    
    # === 判断题型 ===
    # 1. 正误判断题（Gold包含"正确"/"错误"/"一致"/"不一致"/"符合"/"不符合"）
    judge_keywords = ["正确", "错误", "一致", "不一致", "符合", "不符合", "无误", "有误"]
    is_judge = any(kw in gold_str for kw in judge_keywords)
    
    if is_judge:
        # 判断Gold的立场
        if "不正确" in gold_str or "错误" in gold_str or "不符合" in gold_str or "不一致" in gold_str or "有误" in gold_str:
            gold_stand = "negative"
        else:
            gold_stand = "positive"
        
        # 判断回答的立场
        ans_negative = any(kw in ans_str for kw in ["不正确", "错误", "不符合", "不一致", "有误"])
        ans_positive = any(kw in ans_str for kw in ["正确", "无误", "符合", "一致"]) and not ans_negative
        
        # 优先看结论部分
        concl_negative = any(kw in concl_str for kw in ["不正确", "错误", "不符合", "不一致", "有误"])
        concl_positive = any(kw in concl_str for kw in ["正确", "无误", "符合", "一致"]) and not concl_negative
        
        if concl_negative or concl_positive:
            ans_stand = "negative" if concl_negative else "positive"
        elif ans_negative or ans_positive:
            ans_stand = "negative" if ans_negative else "positive"
        else:
            return "UNKNOWN", 0, "无法判断立场"
        
        if ans_stand == gold_stand:
            return "CORRECT", 1.0, f"判断一致（{gold_stand}）"
        else:
            return "INCORRECT", 0, f"判断不一致（gold={gold_stand}, answer={ans_stand}）"
    
    # 2. 选择题（Gold是选项字母）
    gold_choice_match = re.match(r"^([A-D])[、\.\s]", gold_str)
    if gold_choice_match or (len(gold_str) <= 3 and re.match(r"^[A-D]", gold_str)):
        gold_choice = gold_str[0] if gold_str[0] in "ABCD" else None
        if gold_choice:
            # 在回答中找选项
            ans_choices = re.findall(r"[选答案择\s\.]+([A-D])[、\.\s]", ans_str)
            if not ans_choices:
                ans_choices = re.findall(r"([A-D])[、\.]", concl_str)
            
            if ans_choices:
                ans_choice = ans_choices[0]
                if ans_choice == gold_choice:
                    return "CORRECT", 1.0, f"选项一致（{gold_choice}）"
                else:
                    return "INCORRECT", 0, f"选项不一致（gold={gold_choice}, answer={ans_choice}）"
            else:
                return "UNKNOWN", 0.5, "未明确给出选项，需人工复核"
    
    # 3. 计算题（Gold包含数字和单位）
    number_pattern = r"(\d+(?:\.\d+)?)\s*(%|mg/L|mg/m³|mg/Nm³|kg/h|t/a|m³/h|m³|天|元|倍)"
    gold_numbers = re.findall(number_pattern, gold_str)
    
    if gold_numbers and len(gold_numbers) >= 1:
        # 提取回答中的数字
        ans_numbers = re.findall(number_pattern, ans_str)
        if not ans_numbers:
            # 尝试更宽松的匹配
            ans_numbers = [(n, "") for n in re.findall(r"\d+(?:\.\d+)?", concl_str)]
        
        if ans_numbers:
            # 检查关键数值是否匹配（允许10%误差）
            correct_count = 0
            for g_num, g_unit in gold_numbers[:3]:  # 只检查前3个关键数字
                try:
                    g_val = float(g_num)
                    matched = False
                    for a_num, a_unit in ans_numbers:
                        try:
                            a_val = float(a_num)
                            # 单位相同或无单位时比较
                            if g_unit == a_unit or not a_unit or not g_unit:
                                # 允许10%相对误差或0.01绝对误差
                                if g_val > 0:
                                    rel_err = abs(a_val - g_val) / g_val
                                    if rel_err < 0.1 or abs(a_val - g_val) < 0.01:
                                        matched = True
                                        break
                                elif a_val == 0:
                                    matched = True
                                    break
                        except:
                            pass
                    if matched:
                        correct_count += 1
                except:
                    pass
            
            if correct_count == len(gold_numbers[:3]):
                return "CORRECT", 1.0, f"数值匹配（{len(gold_numbers[:3])}个关键数字）"
            elif correct_count > 0:
                return "PARTIAL", correct_count / len(gold_numbers[:3]), f"部分数值匹配（{correct_count}/{len(gold_numbers[:3])}）"
            else:
                return "INCORRECT", 0, "数值不匹配"
    
    # 4. 综合题（包含多个要点）— 关键词匹配
    gold_keywords = re.findall(r"[\u4e00-\u9fff]{2,6}", gold_str)
    # 过滤掉通用词
    stop_words = {"正确", "错误", "符合", "不符合", "一致", "不一致", "根据", "项目", "报告", "分析", "结论", "依据"}
    gold_keywords = [k for k in gold_keywords if k not in stop_words and len(k) >= 2]
    
    if gold_keywords:
        matched = sum(1 for kw in gold_keywords[:10] if kw in ans_str)
        total = min(len(gold_keywords[:10]), 5)
        ratio = matched / total if total > 0 else 0
        
        if ratio >= 0.8:
            return "CORRECT", 1.0, f"关键词匹配率 {ratio:.0%}"
        elif ratio >= 0.5:
            return "PARTIAL", ratio, f"关键词匹配率 {ratio:.0%}"
        else:
            return "INCORRECT", ratio, f"关键词匹配率 {ratio:.0%}"
    
    return "UNKNOWN", 0.5, "无法自动评分，需人工复核"

def main():
    print("=" * 60)
    print("Pilot 16 题自动评分")
    print("=" * 60)
    
    # 加载数据
    questions = load_questions()
    results = load_results()
    
    print(f"\n题目数：{len(questions)}")
    print(f"结果记录：{len(results)} 条")
    
    if not results:
        print("❌ 没有结果数据，请先运行实验")
        return
    
    # 逐题评分
    scored = []
    for r in results:
        qid = r["question_id"]
        q = questions.get(qid, {})
        gold = q.get("gold_answer", "")
        answer = r.get("raw_answer", "")
        
        correctness, score, reason = score_answer(q, gold, answer)
        
        scored.append({
            **r,
            "correctness": correctness,
            "score": score,
            "score_reason": reason,
            "conclusion_extracted": extract_conclusion(answer),
            "needs_human_review": correctness in ("UNKNOWN", "PARTIAL"),
        })
    
    # 生成评分表Excel
    wb = openpyxl.Workbook()
    
    # Sheet1: 逐题评分
    ws1 = wb.active
    ws1.title = "逐题评分"
    headers = [
        "question_id", "EP类别", "task_type", "模型条件", "知识条件",
        "correctness", "score", "评分理由", "是否需人工复核",
        "提取的结论", "Gold答案", "原始回答",
        "input_tokens", "output_tokens", "latency", "status"
    ]
    ws1.append(headers)
    
    for s in scored:
        ws1.append([
            s.get("question_id", ""),
            s.get("ep_category", ""),
            s.get("task_type", ""),
            f"{s.get('model_condition', '')} ({s.get('model_name', '')})",
            s.get("knowledge_condition", ""),
            s.get("correctness", ""),
            s.get("score", 0),
            s.get("score_reason", ""),
            "是" if s.get("needs_human_review") else "否",
            s.get("conclusion_extracted", "")[:300],
            str(s.get("gold_answer", ""))[:300],
            s.get("raw_answer", "")[:500],
            s.get("input_tokens", 0),
            s.get("output_tokens", 0),
            s.get("latency", 0),
            s.get("status", ""),
        ])
    
    # 调整列宽
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 15
    ws1.column_dimensions["E"].width = 20
    ws1.column_dimensions["G"].width = 30
    ws1.column_dimensions["J"].width = 40
    ws1.column_dimensions["K"].width = 40
    ws1.column_dimensions["L"].width = 60
    
    # Sheet2: 汇总统计
    ws2 = wb.create_sheet("汇总统计")
    
    # 按模型×知识条件汇总
    ws2.append(["模型条件", "知识条件", "总题数", "正确", "部分正确", "错误", "不确定", "正确率", "平均分"])
    
    from collections import defaultdict
    groups = defaultdict(list)
    for s in scored:
        key = (s.get("model_condition", ""), s.get("knowledge_condition", ""))
        groups[key].append(s)
    
    for (model, knowledge) in sorted(groups.keys()):
        items = groups[(model, knowledge)]
        total = len(items)
        correct = sum(1 for s in items if s["correctness"] == "CORRECT")
        partial = sum(1 for s in items if s["correctness"] == "PARTIAL")
        incorrect = sum(1 for s in items if s["correctness"] == "INCORRECT")
        unknown = sum(1 for s in items if s["correctness"] == "UNKNOWN")
        avg_score = sum(s["score"] for s in items) / total if total > 0 else 0
        accuracy = correct / total if total > 0 else 0
        
        ws2.append([model, knowledge, total, correct, partial, incorrect, unknown,
                    f"{accuracy:.1%}", f"{avg_score:.3f}"])
    
    ws2.append([])
    
    # 按EP类别汇总
    ws2.append(["EP类别", "总题数", "正确", "正确率", "平均分"])
    ep_groups = defaultdict(list)
    for s in scored:
        ep_groups[s.get("ep_category", "")].append(s)
    
    for ep in sorted(ep_groups.keys()):
        items = ep_groups[ep]
        total = len(items)
        correct = sum(1 for s in items if s["correctness"] == "CORRECT")
        avg_score = sum(s["score"] for s in items) / total if total > 0 else 0
        accuracy = correct / total if total > 0 else 0
        ws2.append([ep, total, correct, f"{accuracy:.1%}", f"{avg_score:.3f}"])
    
    ws2.append([])
    
    # 需人工复核的题
    need_review = [s for s in scored if s.get("needs_human_review")]
    ws2.append(["需人工复核", f"{len(need_review)} 题"])
    for s in need_review[:20]:
        ws2.append(["", f"{s['question_id']} {s['model_condition']}_{s['knowledge_condition']}: {s['score_reason']}"])
    
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 18
    
    wb.save(OUTPUT_XLSX)
    
    # 生成Markdown汇总
    total = len(scored)
    correct = sum(1 for s in scored if s["correctness"] == "CORRECT")
    partial = sum(1 for s in scored if s["correctness"] == "PARTIAL")
    incorrect = sum(1 for s in scored if s["correctness"] == "INCORRECT")
    unknown = sum(1 for s in scored if s["correctness"] == "UNKNOWN")
    avg_score = sum(s["score"] for s in scored) / total if total > 0 else 0
    
    md = f"""# Pilot 16 题评分汇总

- **总调用数**：{total}
- **正确率**：{correct/total:.1%}（{correct}/{total}）
- **平均分**：{avg_score:.3f}

## 正确性分布

| 类别 | 数量 | 占比 |
|------|------|------|
| ✅ 正确 (CORRECT) | {correct} | {correct/total:.1%} |
| ⚠️ 部分正确 (PARTIAL) | {partial} | {partial/total:.1%} |
| ❌ 错误 (INCORRECT) | {incorrect} | {incorrect/total:.1%} |
| ❓ 不确定 (UNKNOWN) | {unknown} | {unknown/total:.1%} |

## 按模型 × 知识条件

| 模型 | 知识条件 | 题数 | 正确率 | 平均分 |
|------|---------|------|--------|--------|
"""
    
    for (model, knowledge) in sorted(groups.keys()):
        items = groups[(model, knowledge)]
        total_g = len(items)
        correct_g = sum(1 for s in items if s["correctness"] == "CORRECT")
        avg_g = sum(s["score"] for s in items) / total_g if total_g > 0 else 0
        md += f"| {model} | {knowledge} | {total_g} | {correct_g/total_g:.1%} | {avg_g:.3f} |\n"
    
    md += f"""
## 需人工复核

共 **{len(need_review)}** 题需人工复核（PARTIAL 或 UNKNOWN）。

详细评分见 `pilot16_scored.xlsx`。
"""
    
    with open(SUMMARY_MD, "w", encoding="utf-8") as f:
        f.write(md)
    
    print(f"\n评分完成！")
    print(f"  总记录：{total}")
    print(f"  正确：{correct} ({correct/total:.1%})")
    print(f"  部分正确：{partial} ({partial/total:.1%})")
    print(f"  错误：{incorrect} ({incorrect/total:.1%})")
    print(f"  不确定：{unknown} ({unknown/total:.1%})")
    print(f"  平均分：{avg_score:.3f}")
    print(f"  需人工复核：{len(need_review)} 题")
    print(f"\n评分表：{OUTPUT_XLSX}")
    print(f"汇总报告：{SUMMARY_MD}")

if __name__ == "__main__":
    main()
