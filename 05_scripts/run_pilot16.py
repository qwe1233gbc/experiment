#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pilot 16 题主实验运行脚本
支持 K1（无知识）和 K3（领域RAG）两种条件
3 个模型：qwen3.8-flash, qwen3.7-max, qwen3.8-max
共 16 题 × 2 知识条件 × 3 模型 = 96 次调用
"""
import os
import sys
import json
import time
import hashlib
import openpyxl
import urllib.request
import urllib.error
import datetime
from pathlib import Path

# ========== 配置 ==========
EXP_DIR = Path(r"E:\实验文件整理_按论文逻辑\实验")
QFILE = EXP_DIR / "02_evaluation_set" / "pilot16_questions.xlsx"
REPORT_DIR = EXP_DIR / "09_input_reports"
RAG_SNAPSHOT = EXP_DIR / "03_knowledge_base" / "pilot16_rag_snapshot.jsonl"
OUTPUT_DIR = EXP_DIR / "07_results"
LOG_DIR = OUTPUT_DIR / "logs"

API_BASE = os.environ.get("API_BASE_URL", "https://one-hub.hycx-gd.cn/v1")
API_KEY = os.environ.get("COMPANY_API_KEY", "")
TEMPERATURE = 0
MAX_TOKENS = 3000

# 模型配置
MODELS = [
    ("M1", "qwen3.8-flash", "轻量版"),
    ("M2", "qwen3.7-max", "上一代旗舰"),
    ("M3", "qwen3.8-max", "当前旗舰"),
]

# 知识条件
KNOWLEDGE_CONDITIONS = [
    ("K1", "无知识（仅报告上下文）"),
    ("K3", "领域RAG"),
]
# K2 稍后补充

OUTPUT_FILE = OUTPUT_DIR / "pilot16_raw_results.jsonl"

os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ========== 日志 ==========
def log(msg):
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{ts} {msg}"
    print(line, flush=True)
    with open(LOG_DIR / "run_pilot16.log", "a", encoding="utf-8") as f:
        f.write(line + "\n")

# ========== 加载数据 ==========
def load_questions():
    wb = openpyxl.load_workbook(QFILE, data_only=True)
    ws = wb["01_题目清单"]
    headers = [c.value for c in ws[1]]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        q = dict(zip(headers, row))
        questions.append({
            "question_id": q["question_id"],
            "ep_category": q["EP类别"],
            "task_type": q.get("task_type", ""),
            "question": q["题目（question）"],
            "project": q.get("项目", ""),
            "gold_answer": q.get("金标答案（gold）", ""),
        })
    return questions

def extract_project_id(question_id, project_field):
    """从question_id或project字段提取项目ID。"""
    if project_field and str(project_field).strip().upper().startswith("PL"):
        return str(project_field).strip().upper()
    
    # 从question_id提取，如 NEW_PL001_invest_ratio -> PL001
    if question_id:
        parts = str(question_id).split('_')
        for p in parts:
            if p.upper().startswith("PL") and len(p) <= 10:
                return p.upper()
    
    return None

def load_report_context(question_id, project_id):
    """加载题目对应的报告上下文。"""
    pid = extract_project_id(question_id, project_id)
    if not pid:
        return "", "no_project"
    
    for f in os.listdir(REPORT_DIR):
        if f.upper().startswith(pid + "_") and f.endswith(".json"):
            with open(os.path.join(REPORT_DIR, f), "r", encoding="utf-8") as fh:
                data = json.load(fh)
            # 提取报告的关键部分作为上下文
            # 简化版：取报告的项目名称、建设内容、产排污环节等
            ctx_parts = []
            if isinstance(data, dict):
                # 尝试提取基本信息
                basic = data.get("基本信息", data.get("项目概况", {}))
                if isinstance(basic, dict):
                    for key in ["项目名称", "建设地点", "建设性质", "总投资", "环保投资", "占地面积"]:
                        val = basic.get(key)
                        if val:
                            ctx_parts.append(f"{key}：{val}")
                
                # 建设内容
                content = data.get("建设内容", data.get("项目组成", ""))
                if content:
                    ctx_parts.append(f"建设内容：{str(content)[:500]}")
                
                # 产排污情况
                pollution = data.get("产排污环节", data.get("污染物排放", ""))
                if pollution:
                    ctx_parts.append(f"产排污情况：{str(pollution)[:800]}")
                
                # 环保措施
                measures = data.get("环保措施", "")
                if measures:
                    ctx_parts.append(f"环保措施：{str(measures)[:500]}")
            
            context = "\n\n".join(ctx_parts)
            if not context:
                # 如果提取不到，就取JSON前2000字符
                context = json.dumps(data, ensure_ascii=False)[:2000]
            
            return context, f
    
    return "", "not_found"

def load_rag_evidence(question_id, top_k=3):
    """加载题目对应的RAG检索结果。"""
    if not RAG_SNAPSHOT.exists():
        return [], ""
    
    hits = []
    with open(RAG_SNAPSHOT, "r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                rec = json.loads(line)
                if rec.get("question_id") == question_id and rec.get("rank", 0) > 0:
                    hits.append(rec)
    
    hits.sort(key=lambda x: x.get("rank", 999))
    top_hits = hits[:top_k]
    
    if not top_hits:
        return [], ""
    
    # 格式化为证据文本
    evidence_parts = []
    for i, hit in enumerate(top_hits, 1):
        source = hit.get("source_file", "未知来源")
        text = hit.get("text", hit.get("child_text", ""))
        evidence_parts.append(
            f"【证据 {i}】来源：{source}\n{text[:1500]}"
        )
    
    evidence_text = "\n\n".join(evidence_parts)
    evidence_hash = hashlib.sha256(evidence_text.encode("utf-8")).hexdigest()
    
    return top_hits, evidence_hash

# ========== Prompt 构造 ==========
SYSTEM_PROMPT = """你是一位专业的环境影响评价审核工程师。请根据提供的环评报告信息和相关法规标准，仔细分析并回答问题。

要求：
1. 先给出明确的结论，再说明理由
2. 所有判断必须基于提供的信息，不得编造
3. 涉及标准的判断，请明确指出依据的标准名称和条款
4. 计算题请列出计算过程
5. 回答要条理清晰，简洁准确"""

def build_prompt(q, report_ctx, knowledge_condition, rag_evidence=None):
    """构造完整的Prompt。"""
    user_msg = f"""## 环评报告信息

{report_ctx if report_ctx else "（报告相关数据已嵌入题目）"}

## 问题

{q['question']}

## 参考资料
"""
    
    if knowledge_condition == "K1":
        user_msg += "（无外部参考资料，请仅根据报告信息和你的专业知识回答）"
    elif knowledge_condition == "K3" and rag_evidence:
        user_msg += "以下是检索到的法规标准参考资料（仅供参考，请结合报告信息判断）：\n\n"
        for i, hit in enumerate(rag_evidence, 1):
            source = hit.get("source_file", "未知")
            text = hit.get("text", "")
            user_msg += f"### 参考资料{i}（来源：{source}）\n{text[:1500]}\n\n"
    
    user_msg += """
## 回答格式要求

请按以下格式回答：

【结论】
（明确给出判断结果，如"符合/不符合"、"正确/错误"、具体数值等）

【分析过程】
（逐条说明分析依据、计算过程、引用的标准条款等）

【依据】
（列出判断所依据的法规标准名称及条款号）"""
    
    return SYSTEM_PROMPT, user_msg

# ========== API 调用 ==========
def call_api(model_name, system_msg, user_msg, retries=3):
    """调用OpenAI兼容接口。"""
    url = f"{API_BASE}/chat/completions"
    body = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_msg},
        ],
        "temperature": TEMPERATURE,
        "max_tokens": MAX_TOKENS,
    }
    data = json.dumps(body, ensure_ascii=False).encode("utf-8")
    
    last_err = None
    for attempt in range(1, retries + 1):
        req = urllib.request.Request(url, data=data, headers={
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json",
        })
        t0 = time.time()
        try:
            with urllib.request.urlopen(req, timeout=300) as resp:
                out = json.loads(resp.read().decode("utf-8"))
            latency = round(time.time() - t0, 2)
            content = out["choices"][0]["message"].get("content") or ""
            usage = out.get("usage", {})
            finish = out["choices"][0].get("finish_reason")
            status = "OK" if content.strip() and finish == "stop" else ("TRUNCATED" if content.strip() else "EMPTY")
            return {
                "content": content,
                "latency": latency,
                "input_tokens": usage.get("prompt_tokens"),
                "output_tokens": usage.get("completion_tokens"),
                "finish_reason": finish,
                "status": status,
            }
        except urllib.error.HTTPError as e:
            last_err = f"HTTP {e.code}: {e.read().decode('utf-8', 'ignore')[:300]}"
        except Exception as e:
            last_err = f"{type(e).__name__}: {str(e)[:300]}"
        
        if attempt < retries:
            wait = 15 * attempt
            log(f"  attempt{attempt} 失败，{wait}s后重试...")
            time.sleep(wait)
    
    return {"content": "", "status": "ERROR", "error": last_err, "latency": None}

# ========== 主函数 ==========
def main():
    log("=" * 60)
    log("Pilot 16 题主实验")
    log("=" * 60)
    
    if not API_KEY:
        log("❌ 缺少 COMPANY_API_KEY 环境变量")
        sys.exit(1)
    
    # 加载题目
    questions = load_questions()
    log(f"加载题目：{len(questions)} 题")
    
    # 统计总调用数
    total_runs = len(questions) * len(MODELS) * len(KNOWLEDGE_CONDITIONS)
    log(f"实验设计：{len(questions)}题 × {len(MODELS)}模型 × {len(KNOWLEDGE_CONDITIONS)}知识条件 = {total_runs} 次调用")
    
    # 断点续跑
    done_ids = set()
    if OUTPUT_FILE.exists():
        with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    rec = json.loads(line)
                    if rec.get("status") in ("OK", "TRUNCATED"):
                        run_id = f"{rec['question_id']}__{rec['model_condition']}__{rec['knowledge_condition']}"
                        done_ids.add(run_id)
        log(f"断点续跑：已完成 {len(done_ids)} 次")
    
    # 逐题运行
    completed = 0
    for q in questions:
        qid = q["question_id"]
        
        # 加载报告上下文
        report_ctx, report_file = load_report_context(qid, q.get("project", ""))
        
        for m_code, m_name, m_desc in MODELS:
            for k_code, k_desc in KNOWLEDGE_CONDITIONS:
                run_id = f"{qid}__{m_code}__{k_code}"
                if run_id in done_ids:
                    continue
                
                # 加载RAG证据
                rag_hits = []
                evidence_hash = ""
                if k_code == "K3":
                    rag_hits, evidence_hash = load_rag_evidence(qid)
                
                # 构造Prompt
                system_msg, user_msg = build_prompt(q, report_ctx, k_code, rag_hits)
                prompt_hash = hashlib.sha256(
                    (system_msg + "\x00" + user_msg).encode("utf-8")
                ).hexdigest()
                
                log(f"▶️  {run_id} ({m_name}, {k_code})")
                
                # 调用API
                result = call_api(m_name, system_msg, user_msg)
                
                # 记录
                record = {
                    "run_id": run_id,
                    "question_id": qid,
                    "ep_category": q["ep_category"],
                    "task_type": q["task_type"],
                    "project": q.get("project", ""),
                    "model_condition": m_code,
                    "model_name": m_name,
                    "knowledge_condition": k_code,
                    "temperature": TEMPERATURE,
                    "max_tokens": MAX_TOKENS,
                    "report_file": report_file,
                    "report_context_hash": hashlib.sha256(report_ctx.encode("utf-8")).hexdigest() if report_ctx else "",
                    "evidence_hash": evidence_hash,
                    "prompt_hash": prompt_hash,
                    "gold_answer": q["gold_answer"],
                    "input_tokens": result.get("input_tokens"),
                    "output_tokens": result.get("output_tokens"),
                    "latency": result.get("latency"),
                    "status": result.get("status"),
                    "finish_reason": result.get("finish_reason"),
                    "error": result.get("error"),
                    "raw_answer": result.get("content", ""),
                }
                
                with open(OUTPUT_FILE, "a", encoding="utf-8") as f:
                    f.write(json.dumps(record, ensure_ascii=False) + "\n")
                
                status = result.get("status", "?")
                tokens = result.get("output_tokens", "?")
                latency = result.get("latency", "?")
                log(f"  ✅ {status} | {tokens} tokens | {latency}s")
                completed += 1
    
    # 汇总
    log(f"\n{'=' * 60}")
    log(f"完成！本轮新增 {completed} 次调用")
    
    # 统计最终结果
    all_records = []
    if OUTPUT_FILE.exists():
        with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    all_records.append(json.loads(line))
    
    ok = [r for r in all_records if r["status"] == "OK"]
    failed = [r for r in all_records if r["status"] == "ERROR"]
    log(f"总计：{len(all_records)} 条，成功：{len(ok)}，失败：{len(failed)}")
    
    if failed:
        log(f"失败列表：")
        for r in failed:
            log(f"  {r['run_id']}: {r.get('error', '?')[:100]}")
    
    log(f"\n输出文件：{OUTPUT_FILE}")

if __name__ == "__main__":
    main()
