"""重新生成PL001检查Excel — 明确标注模型回复 vs 标准答案。"""
import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO = Path(r"E:\华南理工项目\环评文件汇总\01_GitHub项目与研究文档\eia-openclaw-sync-chen2026\10_消融实验设计")
QA_EXCEL = Path(r"c:\Users\ylx\.trae-cn\attachments\6a7440dbc2e5f7d2fcd31f9a\92a518a4-420c-4a47-a4e9-3f8d906a7654_bdd4d9a4-8da3-4d51-acfc-7c51d714c6fa_四大类问答对_仅保留正确和无误 v01.xlsx")

GROUP_DIRS = {
    "A": REPO / r"06_Qwen3.8Max_A组_20260806\run_20260806_222338",
    "B": REPO / r"07_Qwen3.8Max_B组_20260807\run_20260807_215254",
    "C": REPO / r"05_Qwen3.8Max_CD重跑_20260806\run_20260806_164352",
    "D": REPO / r"05_Qwen3.8Max_CD重跑_20260806\run_20260806_164352",
}

GROUP_CONFIGS = {
    "A": "LLM only (基线)",
    "B": "LLM + RAG (法规库)",
    "C": "LLM + Skill (技能库)",
    "D": "LLM + RAG + Skill (完整版)",
}

QUESTIONS = [
    ("PL001_V01_Q01", "国民经济行业分类", "请结合项目产品、原辅材料和生产工艺，判断该报告行业类别是否基本合理。"),
    ("PL001_EnvQuality_Q01", "环境质量数据引用", "请根据报告引用的环境质量公报及现状数据，判断大气和地表水环境质量现状数据引用是否准确。"),
    ("PL001_Emission_固体", "固体废物控制标准", "请根据项目产生的固体废物和危险废物类型，判断固体废物控制标准选取是否合理。要求总结一般工业固体废物和危险废物执行标准的标准名称及编号，分析是否缺少必要标准条目。"),
]

OUTPUT = Path(r"C:\Users\ylx\AppData\Roaming\TRAE SOLO CN\ModularData\ai-agent\work-mode-projects\6a789bcfc2e5f7d2fcd33cd3\实验文件整理_按论文逻辑\PL001_ABCD组_模型回复与标准答案对比.xlsx")

# ─── Styles ───
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2B6CB0")
LABEL_FONT = Font(name="微软雅黑", bold=True, size=10)
LABEL_FILL = PatternFill("solid", fgColor="EDF2F7")
BODY_FONT = Font(name="微软雅黑", size=10)
REF_FONT = Font(name="微软雅黑", size=10, bold=True, color="22543D")
REF_FILL = PatternFill("solid", fgColor="C6F6D5")
GROUP_FILLS = {
    "A": PatternFill("solid", fgColor="FFF5F5"),
    "B": PatternFill("solid", fgColor="F0FFF4"),
    "C": PatternFill("solid", fgColor="FFFFF0"),
    "D": PatternFill("solid", fgColor="F0F4FF"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E0"),
    right=Side(style="thin", color="CBD5E0"),
    top=Side(style="thin", color="CBD5E0"),
    bottom=Side(style="thin", color="CBD5E0"),
)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def load_reference_answers():
    """从上传的Excel读取标准答案。"""
    df = pd.read_excel(str(QA_EXCEL))
    refs = {}
    for _, row in df.iterrows():
        qid = str(row.get("question_id", ""))
        if qid.startswith("PL001"):
            refs[qid] = {
                "answer": str(row.get("answer", "")),
                "evidence": str(row.get("evidence", "")),
                "source_basis": str(row.get("source_basis", "")),
                "manual_judgment": str(row.get("人工判断", "")),
                "manual_note": str(row.get("人工备注", "")),
                "polished_answer": str(row.get("润色后答案", "")),
                "ai_note": str(row.get("AI标注备注", "")),
            }
    return refs


def load_response(group, qid):
    base = GROUP_DIRS[group]
    p = base / "parsed_outputs" / f"{qid}.json"
    if not p.exists():
        p = base / "parsed_outputs" / group / f"{qid}.json"
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return {}


def load_raw_text(group, qid):
    base = GROUP_DIRS[group]
    p = base / "raw_responses" / group / f"{qid}.json"
    if p.exists():
        data = json.loads(p.read_text(encoding="utf-8"))
        return data.get("output_text", "")
    return ""


def format_evidence(items):
    if not items:
        return "(无)"
    lines = []
    for i, e in enumerate(items, 1):
        lines.append(f"[证据{i}] {e.get('fact','')}\n  位置: {e.get('source_location','')}\n  摘录: {e.get('source_excerpt','')}")
    return "\n".join(lines)


def format_checks(items):
    if not items:
        return "(无)"
    lines = []
    for i, c in enumerate(items, 1):
        lines.append(f"[检查{i}] {c.get('check_name','')}\n  结果: {c.get('result','')}\n  说明: {c.get('explanation','')}")
    return "\n".join(lines)


def format_calcs(items):
    if not items:
        return "(无)"
    lines = []
    for i, c in enumerate(items, 1):
        lines.append(f"[计算{i}] {c.get('formula','')}\n  输入: {json.dumps(c.get('inputs',{}), ensure_ascii=False)}\n  结果: {c.get('result','')} {c.get('unit','')}")
    return "\n".join(lines)


def format_issues(items):
    if not items:
        return "(无问题发现)"
    return "\n".join(f"[问题{i+1}] {v}" for i, v in enumerate(items))


def format_refs(items):
    if not items:
        return "(无)"
    lines = []
    for i, r in enumerate(items, 1):
        lines.append(f"[引用{i}] {r.get('reference_id','')} - {r.get('title','')}\n  条款: {r.get('clause_or_table','')}\n  支持文本: {r.get('supporting_text','')}")
    return "\n".join(lines)


def format_missing(items):
    if not items:
        return "(无缺失)"
    return "\n".join(f"[缺失{i+1}] {v}" for i, v in enumerate(items))


def create_excel():
    refs = load_reference_answers()
    print(f"标准答案加载: {list(refs.keys())}")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ═══ 汇总对比sheet ═══
    ws = wb.create_sheet("汇总对比")
    
    # 表头
    headers = [
        "行类型", "题目编号", "审核类别", "问题",
        "组别", "配置/来源", "判定结论", "详细分析",
        "内部审核状态", "内部审核结论",
        "外部验证状态", "外部验证结论",
        "应答模式", "幻觉标记", "需人工复核",
        "人工判断", "人工备注"
    ]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER
    
    for qid, category, question in QUESTIONS:
        ref = refs.get(qid, {})
        
        # 先输出标准答案行（绿色标注）
        ref_row = [
            "★ 标准答案（人工核定）", qid, category, question,
            "—", "四大类问答对Excel（人工核定）",
            ref.get("manual_judgment", ""),
            ref.get("polished_answer", ref.get("answer", "")),
            "—", "—", "—", "—", "—", "—", "—",
            ref.get("manual_judgment", ""),
            ref.get("manual_note", ""),
        ]
        ws.append(ref_row)
        r = ws.max_row
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=r, column=col)
            cell.font = REF_FONT
            cell.fill = REF_FILL
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
        
        # 然后输出ABCD四组模型回复
        for group in ["A", "B", "C", "D"]:
            oj = load_response(group, qid)
            ri = oj.get("report_internal_result", {})
            ev = oj.get("external_validation_result", {})
            fa = oj.get("final_answer", {})
            sf = oj.get("safety_flags", {})
            
            model_row = [
                f"● qwen3.8-max {group}组回复", qid, category, question,
                group, GROUP_CONFIGS[group],
                fa.get("judgement", ""),
                fa.get("analysis", ""),
                ri.get("status", ""),
                ri.get("conclusion", ""),
                ev.get("status", ""),
                ev.get("conclusion", ""),
                oj.get("answer_mode", ""),
                "是 ⚠️" if str(sf.get("used_unprovided_external_knowledge", "")).lower() == "true" else "否",
                "是" if str(sf.get("manual_review_needed", "")).lower() == "true" else "否",
                "—", "—",
            ]
            ws.append(model_row)
            r = ws.max_row
            for col in range(1, len(headers)+1):
                cell = ws.cell(row=r, column=col)
                cell.font = BODY_FONT
                cell.alignment = WRAP_TOP
                cell.border = THIN_BORDER
                cell.fill = GROUP_FILLS[group]
    
    # 列宽
    widths = [20, 18, 14, 40, 6, 22, 12, 50, 14, 40, 18, 40, 16, 10, 10, 12, 20]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64+i) if i <= 26 else "A" + chr(64+i-26)
        ws.column_dimensions[col_letter].width = w
    
    ws.freeze_panes = "A2"
    
    # ═══ 每题详细对比sheet ═══
    for qid, category, question in QUESTIONS:
        ws2 = wb.create_sheet(qid[:30])
        ref = refs.get(qid, {})
        
        # 标题
        ws2.merge_cells("A1:F1")
        ws2["A1"] = f"题目: {qid}  |  审核类别: {category}"
        ws2["A1"].font = Font(name="微软雅黑", bold=True, size=13, color="1A365D")
        ws2["A1"].alignment = WRAP_CENTER
        
        ws2.merge_cells("A2:F2")
        ws2["A2"] = f"问题: {question}"
        ws2["A2"].font = Font(name="微软雅黑", size=10, color="4A5568")
        ws2["A2"].alignment = WRAP_TOP
        
        # 表头
        row = 4
        headers2 = ["对比维度", "★ 标准答案(人工核定)", "A组 (LLM only)", "B组 (LLM+RAG)", "C组 (LLM+Skill)", "D组 (LLM+RAG+Skill)"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP_CENTER
            cell.border = THIN_BORDER
        
        # 加载四组数据
        all_data = {}
        for group in ["A", "B", "C", "D"]:
            all_data[group] = load_response(group, qid)
        
        # 标准答案提取器
        ref_extractors = {
            "判定结论": lambda: ref.get("manual_judgment", ""),
            "标准答案(完整)": lambda: ref.get("polished_answer", ref.get("answer", "")),
            "标准答案(原始)": lambda: ref.get("answer", ""),
            "证据说明": lambda: ref.get("evidence", ""),
            "来源依据": lambda: ref.get("source_basis", ""),
            "AI标注备注": lambda: ref.get("ai_note", ""),
        }
        
        # 模型回复提取器
        model_dims = [
            ("判定 (judgement)", lambda oj: oj.get("final_answer", {}).get("judgement", "")),
            ("分析 (analysis)", lambda oj: oj.get("final_answer", {}).get("analysis", "")),
            ("范围说明 (scope_note)", lambda oj: oj.get("final_answer", {}).get("scope_note", "")),
            ("内部审核状态", lambda oj: oj.get("report_internal_result", {}).get("status", "")),
            ("内部审核结论", lambda oj: oj.get("report_internal_result", {}).get("conclusion", "")),
            ("证据 (evidence)", lambda oj: format_evidence(oj.get("report_internal_result", {}).get("evidence", []))),
            ("检查项 (checks)", lambda oj: format_checks(oj.get("report_internal_result", {}).get("checks", []))),
            ("计算 (calculations)", lambda oj: format_calcs(oj.get("report_internal_result", {}).get("calculations", []))),
            ("发现问题 (issues)", lambda oj: format_issues(oj.get("report_internal_result", {}).get("issues", []))),
            ("外部验证状态", lambda oj: oj.get("external_validation_result", {}).get("status", "")),
            ("外部验证结论", lambda oj: oj.get("external_validation_result", {}).get("conclusion", "")),
            ("引用法规 (references_used)", lambda oj: format_refs(oj.get("external_validation_result", {}).get("references_used", []))),
            ("缺失依据 (missing_references)", lambda oj: format_missing(oj.get("external_validation_result", {}).get("missing_references", []))),
            ("应答模式 (answer_mode)", lambda oj: oj.get("answer_mode", "")),
            ("幻觉标记", lambda oj: "是 ⚠️" if str(oj.get("safety_flags", {}).get("used_unprovided_external_knowledge", "")).lower() == "true" else "否"),
            ("整体弃权", lambda oj: "是" if str(oj.get("safety_flags", {}).get("whole_task_abstention", "")).lower() == "true" else "否"),
            ("需人工复核", lambda oj: "是" if str(oj.get("safety_flags", {}).get("manual_review_needed", "")).lower() == "true" else "否"),
            ("技能ID (skill_id)", lambda oj: oj.get("skill_id", "") or "(无)"),
        ]
        
        # 标准答案行
        for dim_name, extractor in ref_extractors.items():
            row += 1
            ws2.cell(row=row, column=1, value=dim_name)
            ws2.cell(row=row, column=1).font = LABEL_FONT
            ws2.cell(row=row, column=1).fill = LABEL_FILL
            ws2.cell(row=row, column=1).alignment = WRAP_TOP
            ws2.cell(row=row, column=1).border = THIN_BORDER
            
            val = extractor()
            cell = ws2.cell(row=row, column=2, value=val)
            cell.font = REF_FONT
            cell.fill = REF_FILL
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
            
            # 模型回复列留空（与标准答案不在同行）
            for col in range(3, 7):
                ws2.cell(row=row, column=col, value="").border = THIN_BORDER
        
        # 分隔行
        row += 1
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws2.cell(row=row, column=1, value="▼ 以下为qwen3.8-max模型回复（A/B/C/D四组对比）▼")
        ws2.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=10, color="C53030")
        ws2.cell(row=row, column=1).fill = PatternFill("solid", fgColor="FED7D7")
        ws2.cell(row=row, column=1).alignment = WRAP_CENTER
        
        # 模型回复行
        for dim_name, extractor in model_dims:
            row += 1
            ws2.cell(row=row, column=1, value=dim_name)
            ws2.cell(row=row, column=1).font = LABEL_FONT
            ws2.cell(row=row, column=1).fill = LABEL_FILL
            ws2.cell(row=row, column=1).alignment = WRAP_TOP
            ws2.cell(row=row, column=1).border = THIN_BORDER
            
            # 标准答案列显示"—"
            ws2.cell(row=row, column=2, value="—")
            ws2.cell(row=row, column=2).font = Font(name="微软雅黑", size=9, color="A0AEC0")
            ws2.cell(row=row, column=2).alignment = WRAP_TOP
            ws2.cell(row=row, column=2).border = THIN_BORDER
            
            for col, group in enumerate(["A", "B", "C", "D"], 3):
                val = extractor(all_data[group])
                cell = ws2.cell(row=row, column=col, value=val)
                cell.font = BODY_FONT
                cell.alignment = WRAP_TOP
                cell.border = THIN_BORDER
                cell.fill = GROUP_FILLS[group]
        
        # 列宽
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 55
        ws2.column_dimensions["C"].width = 50
        ws2.column_dimensions["D"].width = 50
        ws2.column_dimensions["E"].width = 50
        ws2.column_dimensions["F"].width = 50
        
        ws2.freeze_panes = "C5"
    
    # ═══ 原始回复sheet ═══
    ws_raw = wb.create_sheet("原始output_text")
    raw_headers = ["行类型", "题目编号", "组别/来源", "配置", "判定", "原始回复(output_text)"]
    ws_raw.append(raw_headers)
    for col in range(1, len(raw_headers)+1):
        cell = ws_raw.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER
    
    # 标准答案行
    for qid, category, question in QUESTIONS:
        ref = refs.get(qid, {})
        ws_raw.append([
            "★ 标准答案", qid, "人工核定", "四大类问答对Excel",
            ref.get("manual_judgment", ""),
            ref.get("polished_answer", ref.get("answer", ""))
        ])
        r = ws_raw.max_row
        for col in range(1, len(raw_headers)+1):
            cell = ws_raw.cell(row=r, column=col)
            cell.font = REF_FONT
            cell.fill = REF_FILL
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
    
    # 模型回复行
    for qid, category, question in QUESTIONS:
        for group in ["A", "B", "C", "D"]:
            oj = load_response(group, qid)
            raw = load_raw_text(group, qid)
            judgement = oj.get("final_answer", {}).get("judgement", "")
            ws_raw.append([
                f"● {group}组模型回复", qid, group, GROUP_CONFIGS[group],
                judgement, raw
            ])
            r = ws_raw.max_row
            for col in range(1, len(raw_headers)+1):
                cell = ws_raw.cell(row=r, column=col)
                cell.font = BODY_FONT
                cell.alignment = WRAP_TOP
                cell.border = THIN_BORDER
                cell.fill = GROUP_FILLS[group]
    
    ws_raw.column_dimensions["A"].width = 16
    ws_raw.column_dimensions["B"].width = 20
    ws_raw.column_dimensions["C"].width = 10
    ws_raw.column_dimensions["D"].width = 22
    ws_raw.column_dimensions["E"].width = 10
    ws_raw.column_dimensions["F"].width = 120
    
    ws_raw.freeze_panes = "A2"
    
    wb.save(str(OUTPUT))
    print(f"Excel已保存: {OUTPUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    create_excel()
